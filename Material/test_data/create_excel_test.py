#!/usr/bin/env python3
"""
Script per creare file Excel di test per Wire Editor Unified
Eseguire sul sistema dove openpyxl è installato
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ Openpyxl non installato!")
    print("   Installare con: pip install openpyxl")
    exit(1)

# Crea workbook
wb = Workbook()
ws = wb.active
ws.title = "Coordinate"

# Header con stile
headers = ['ID', 'X (m)', 'Y (m)', 'Z (m)', 'Livello', 'Descrizione']
ws.append(headers)

# Stile header
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Dati di esempio - Edificio 3 piani
data = [
    [1, 0.0000, 0.0000, 0.000, "PT", "Punto di riferimento"],
    [2, 5.5000, 0.0000, 0.000, "PT", "Angolo Nord-Est"],
    [3, 5.5000, 8.2500, 0.000, "PT", "Angolo Sud-Est"],
    [4, 0.0000, 8.2500, 0.000, "PT", "Angolo Sud-Ovest"],
    [5, 2.7500, 0.0000, 3.200, "P1", "Filo piano 1 - Centro Nord"],
    [6, 2.7500, 8.2500, 3.200, "P1", "Filo piano 1 - Centro Sud"],
    [7, 0.0000, 4.1250, 3.200, "P1", "Filo piano 1 - Lato Ovest"],
    [8, 5.5000, 4.1250, 3.200, "P1", "Filo piano 1 - Lato Est"],
    [9, 2.7500, 4.1250, 6.400, "P2", "Filo piano 2 - Centro"],
    [10, 1.3750, 2.0625, 6.400, "P2", "Filo piano 2 - Quadrante NO"],
    [11, 4.1250, 2.0625, 6.400, "P2", "Filo piano 2 - Quadrante NE"],
    [12, 4.1250, 6.1875, 6.400, "P2", "Filo piano 2 - Quadrante SE"],
    [13, 1.3750, 6.1875, 6.400, "P2", "Filo piano 2 - Quadrante SO"],
    [14, 2.7500, 4.1250, 9.600, "COPERTURA", "Colmo copertura"],
    [15, 0.5000, 1.0000, 9.600, "COPERTURA", "Filo di gronda Nord-Ovest"],
]

for row in data:
    ws.append(row)

# Formatta numeri
for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=2, max_col=4):
    for cell in row:
        cell.number_format = '0.0000' if cell.column in [2, 3] else '0.000'
        cell.alignment = Alignment(horizontal="right")

# Auto-width colonne
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

# Salva
output_file = 'test_coordinates.xlsx'
wb.save(output_file)
print(f"✅ File Excel creato: {output_file}")
print(f"   15 punti distribuiti su 4 livelli (PT, P1, P2, COPERTURA)")
print(f"   Pronto per import in Wire Editor Unified!")
