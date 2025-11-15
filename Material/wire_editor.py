import sys
import json
import math
import csv
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QDoubleSpinBox, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QGroupBox, QGridLayout, QFileDialog, QMessageBox,
    QHeaderView, QAbstractItemView, QDialog, QTextEdit, QDialogButtonBox,
    QInputDialog, QComboBox, QSpinBox, QCheckBox
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont

# Matplotlib imports (opzionali)
try:
    import matplotlib
    matplotlib.use('Qt5Agg')
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# Openpyxl imports (opzionali)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Reportlab imports (opzionali)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Ezdxf imports (opzionali - per import DXF avanzato)
try:
    import ezdxf
    EZDXF_AVAILABLE = True
except ImportError:
    EZDXF_AVAILABLE = False


class Command:
    """Classe base per comandi Undo/Redo"""
    def execute(self, model):
        """Esegue il comando"""
        raise NotImplementedError

    def undo(self, model):
        """Annulla il comando"""
        raise NotImplementedError


class AddPointCommand(Command):
    """Comando per aggiungere un punto"""
    def __init__(self, x, y, z, description, level):
        self.x = x
        self.y = y
        self.z = z
        self.description = description
        self.level = level
        self.point_id = None

    def execute(self, model):
        self.point_id = model._add_point_internal(self.x, self.y, self.z, self.description, self.level)
        return self.point_id

    def undo(self, model):
        if self.point_id:
            model._delete_point_internal(self.point_id)


class DeletePointCommand(Command):
    """Comando per eliminare un punto"""
    def __init__(self, point_id):
        self.point_id = point_id
        self.saved_point = None

    def execute(self, model):
        if self.point_id in model.points:
            # Salva punto per undo
            p = model.points[self.point_id]
            self.saved_point = (p.id, p.x, p.y, p.z, p.description, p.level)
            model._delete_point_internal(self.point_id)
            return True
        return False

    def undo(self, model):
        if self.saved_point:
            pid, x, y, z, desc, level = self.saved_point
            # Ripristina con ID originale
            model._restore_point_internal(pid, x, y, z, desc, level)


class UpdatePointCommand(Command):
    """Comando per aggiornare un punto"""
    def __init__(self, point_id, **kwargs):
        self.point_id = point_id
        self.new_values = kwargs
        self.old_values = {}

    def execute(self, model):
        if self.point_id in model.points:
            point = model.points[self.point_id]
            # Salva valori vecchi
            for key in self.new_values.keys():
                if hasattr(point, key):
                    self.old_values[key] = getattr(point, key)
            # Applica nuovi valori
            model._update_point_internal(self.point_id, **self.new_values)
            return True
        return False

    def undo(self, model):
        if self.old_values:
            model._update_point_internal(self.point_id, **self.old_values)


class TranslateAllCommand(Command):
    """Comando per traslare tutti i punti"""
    def __init__(self, dx, dy, dz):
        self.dx = dx
        self.dy = dy
        self.dz = dz

    def execute(self, model):
        model._translate_all_internal(self.dx, self.dy, self.dz)

    def undo(self, model):
        model._translate_all_internal(-self.dx, -self.dy, -self.dz)


class ClearAllCommand(Command):
    """Comando per pulire tutti i punti"""
    def __init__(self):
        self.saved_points = []
        self.saved_next_id = 1

    def execute(self, model):
        # Salva tutti i punti
        self.saved_points = [(p.id, p.x, p.y, p.z, p.description, p.level)
                            for p in model.points.values()]
        self.saved_next_id = model.next_id
        model._clear_internal()

    def undo(self, model):
        # Ripristina tutti i punti
        for pid, x, y, z, desc, level in self.saved_points:
            model._restore_point_internal(pid, x, y, z, desc, level)
        model.next_id = self.saved_next_id


class Point:
    """Punto con coordinate spaziali"""

    def __init__(self, point_id, x, y, z=0.0, description="", level=""):
        self.id = point_id
        self.x = x
        self.y = y
        self.z = z
        self.description = description
        self.level = level  # Livello/Piano (es: "PT", "P1", "P2")

    def to_dict(self):
        return {
            'id': self.id,
            'x': self.x,
            'y': self.y,
            'z': self.z,
            'description': self.description,
            'level': self.level
        }

    @classmethod
    def from_dict(cls, data):
        return cls(
            data['id'],
            data['x'],
            data['y'],
            data.get('z', 0.0),
            data.get('description', ''),
            data.get('level', '')
        )


class CoordinateModel:
    """Modello dati per coordinate fili fissi"""

    def __init__(self):
        self.points = {}
        self.next_id = 1
        self.metadata = {
            "author": "Arch. Michelangelo Bartolotta",
            "registration": "Ordine Architetti Agrigento n. 1557",
            "project_type": "Coordinate Fili Fissi",
            "created": datetime.now().isoformat(),
            "modified": datetime.now().isoformat(),
            "project_name": "",
            "location": "",
            "notes": ""
        }
        # Undo/Redo system
        self.undo_stack = []
        self.redo_stack = []
        self.max_undo_levels = 50  # Limita memoria
    
    # === Metodi interni (usati dai comandi) ===

    def _add_point_internal(self, x, y, z=0.0, description="", level=""):
        """Metodo interno per aggiungere punto (usato da Command)"""
        if not description:
            description = f"Punto {self.next_id}"
        point = Point(self.next_id, x, y, z, description, level)
        self.points[self.next_id] = point
        point_id = self.next_id
        self.next_id += 1
        self._mark_modified()
        return point_id

    def _delete_point_internal(self, point_id):
        """Metodo interno per eliminare punto (usato da Command)"""
        if point_id in self.points:
            del self.points[point_id]
            self._mark_modified()

    def _update_point_internal(self, point_id, **kwargs):
        """Metodo interno per aggiornare punto (usato da Command)"""
        if point_id in self.points:
            point = self.points[point_id]
            for key, value in kwargs.items():
                if hasattr(point, key):
                    setattr(point, key, value)
            self._mark_modified()

    def _restore_point_internal(self, point_id, x, y, z, description, level):
        """Metodo interno per ripristinare punto con ID specifico"""
        point = Point(point_id, x, y, z, description, level)
        self.points[point_id] = point
        if point_id >= self.next_id:
            self.next_id = point_id + 1
        self._mark_modified()

    def _translate_all_internal(self, dx, dy, dz):
        """Metodo interno per traslare tutti i punti"""
        for point in self.points.values():
            point.x += dx
            point.y += dy
            point.z += dz
        self._mark_modified()

    def _clear_internal(self):
        """Metodo interno per pulire tutti i punti"""
        self.points.clear()
        self.next_id = 1
        self._mark_modified()

    # === Sistema Undo/Redo ===

    def execute_command(self, command):
        """Esegue un comando e lo aggiunge allo stack undo"""
        result = command.execute(self)
        self.undo_stack.append(command)
        # Limita dimensione stack
        if len(self.undo_stack) > self.max_undo_levels:
            self.undo_stack.pop(0)
        # Pulisce redo stack quando si esegue nuovo comando
        self.redo_stack.clear()
        return result

    def undo(self):
        """Annulla l'ultima operazione"""
        if not self.undo_stack:
            return False
        command = self.undo_stack.pop()
        command.undo(self)
        self.redo_stack.append(command)
        self._mark_modified()
        return True

    def redo(self):
        """Ripristina l'ultima operazione annullata"""
        if not self.redo_stack:
            return False
        command = self.redo_stack.pop()
        command.execute(self)
        self.undo_stack.append(command)
        self._mark_modified()
        return True

    def can_undo(self):
        """Verifica se è possibile annullare"""
        return len(self.undo_stack) > 0

    def can_redo(self):
        """Verifica se è possibile ripristinare"""
        return len(self.redo_stack) > 0

    def clear_undo_history(self):
        """Pulisce la cronologia undo/redo"""
        self.undo_stack.clear()
        self.redo_stack.clear()

    # === Metodi pubblici (wrapper con comandi) ===

    def add_point(self, x, y, z=0.0, description="", level=""):
        """Aggiunge un punto (con undo support)"""
        cmd = AddPointCommand(x, y, z, description, level)
        return self.execute_command(cmd)
    
    def update_point(self, point_id, **kwargs):
        """Aggiorna un punto (con undo support)"""
        if point_id in self.points:
            cmd = UpdatePointCommand(point_id, **kwargs)
            return self.execute_command(cmd)
        return False

    def delete_point(self, point_id):
        """Elimina un punto (con undo support)"""
        if point_id in self.points:
            cmd = DeletePointCommand(point_id)
            return self.execute_command(cmd)
        return False
    
    def get_point_list(self):
        """Restituisce lista ordinata di punti"""
        return sorted(self.points.values(), key=lambda p: p.id)
    
    def clear(self):
        """Pulisce tutti i punti (con undo support)"""
        cmd = ClearAllCommand()
        self.execute_command(cmd)
    
    def get_bounds(self):
        """Calcola i limiti delle coordinate"""
        if not self.points:
            return None

        points = list(self.points.values())
        return {
            'x_min': min(p.x for p in points),
            'x_max': max(p.x for p in points),
            'y_min': min(p.y for p in points),
            'y_max': max(p.y for p in points),
            'z_min': min(p.z for p in points),
            'z_max': max(p.z for p in points)
        }

    def duplicate_point(self, point_id, offset_x=0.0, offset_y=0.0, offset_z=0.0):
        """Duplica un punto con offset opzionale"""
        if point_id in self.points:
            original = self.points[point_id]
            new_desc = f"{original.description} (copia)" if original.description else f"Punto {self.next_id}"
            return self.add_point(
                original.x + offset_x,
                original.y + offset_y,
                original.z + offset_z,
                new_desc
            )
        return None

    def translate_all(self, dx, dy, dz):
        """Trasla tutti i punti di dx, dy, dz (con undo support)"""
        cmd = TranslateAllCommand(dx, dy, dz)
        self.execute_command(cmd)

    def find_duplicates(self, tolerance=0.001):
        """Trova punti duplicati entro una tolleranza"""
        duplicates = []
        points_list = list(self.points.values())

        for i, p1 in enumerate(points_list):
            for p2 in points_list[i+1:]:
                dist = math.sqrt((p1.x - p2.x)**2 + (p1.y - p2.y)**2 + (p1.z - p2.z)**2)
                if dist < tolerance:
                    duplicates.append((p1.id, p2.id, dist))

        return duplicates

    def calculate_distance(self, id1, id2):
        """Calcola distanza tra due punti"""
        if id1 in self.points and id2 in self.points:
            p1 = self.points[id1]
            p2 = self.points[id2]
            dx = p2.x - p1.x
            dy = p2.y - p1.y
            dz = p2.z - p1.z
            dist_3d = math.sqrt(dx**2 + dy**2 + dz**2)
            dist_2d = math.sqrt(dx**2 + dy**2)
            return {
                'distance_3d': dist_3d,
                'distance_2d': dist_2d,
                'dx': dx,
                'dy': dy,
                'dz': dz
            }
        return None

    def get_levels(self):
        """Restituisce lista livelli unici ordinati"""
        levels = set(p.level for p in self.points.values() if p.level)
        return sorted(levels)

    def get_points_by_level(self, level):
        """Restituisce punti di un livello specifico"""
        return [p for p in self.points.values() if p.level == level]

    def copy_level(self, source_level, target_level, offset_z=0.0):
        """Copia tutti i punti di un livello su un altro livello"""
        source_points = self.get_points_by_level(source_level)
        copied = 0
        for p in source_points:
            self.add_point(p.x, p.y, p.z + offset_z, p.description, target_level)
            copied += 1
        return copied

    def assign_level_to_points(self, point_ids, level):
        """Assegna livello a lista di punti"""
        for pid in point_ids:
            if pid in self.points:
                self.points[pid].level = level
        self._mark_modified()

    def select_points_by_area(self, x_min, x_max, y_min, y_max):
        """Seleziona punti in area rettangolare"""
        selected = []
        for p in self.points.values():
            if x_min <= p.x <= x_max and y_min <= p.y <= y_max:
                selected.append(p.id)
        return selected

    def select_points_by_z(self, z_value, tolerance=0.001):
        """Seleziona punti a quota Z (con tolleranza)"""
        selected = []
        for p in self.points.values():
            if abs(p.z - z_value) < tolerance:
                selected.append(p.id)
        return selected

    def translate_points(self, point_ids, dx, dy, dz):
        """Trasla solo punti selezionati"""
        for pid in point_ids:
            if pid in self.points:
                self.points[pid].x += dx
                self.points[pid].y += dy
                self.points[pid].z += dz
        self._mark_modified()

    def delete_points(self, point_ids):
        """Elimina lista di punti"""
        for pid in point_ids:
            if pid in self.points:
                del self.points[pid]
        self._mark_modified()

    def mirror_points(self, point_ids, axis='x', value=0.0):
        """Specchia punti rispetto ad asse"""
        for pid in point_ids:
            if pid in self.points:
                p = self.points[pid]
                if axis == 'x':
                    p.y = 2 * value - p.y
                elif axis == 'y':
                    p.x = 2 * value - p.x
                elif axis == 'z':
                    p.z = 2 * value - p.z
        self._mark_modified()

    def snap_to_grid(self, point_ids, grid_size=0.5):
        """Arrotonda coordinate a griglia"""
        for pid in point_ids:
            if pid in self.points:
                p = self.points[pid]
                p.x = round(p.x / grid_size) * grid_size
                p.y = round(p.y / grid_size) * grid_size
                p.z = round(p.z / grid_size) * grid_size
        self._mark_modified()

    def generate_grid(self, x_start, x_end, x_count, y_start, y_end, y_count, z, level="", prefix=""):
        """Genera griglia regolare di punti"""
        generated = 0
        x_coords = [x_start + i * (x_end - x_start) / (x_count - 1) for i in range(x_count)] if x_count > 1 else [x_start]
        y_coords = [y_start + i * (y_end - y_start) / (y_count - 1) for i in range(y_count)] if y_count > 1 else [y_start]

        # Lettere per assi X: A, B, C...
        x_labels = [chr(65 + i) for i in range(min(x_count, 26))]
        # Numeri per assi Y: 1, 2, 3...
        y_labels = [str(i + 1) for i in range(y_count)]

        for i, x in enumerate(x_coords):
            for j, y in enumerate(y_coords):
                if i < len(x_labels) and j < len(y_labels):
                    desc = f"{prefix}{x_labels[i]}{y_labels[j]}"
                    self.add_point(x, y, z, desc, level)
                    generated += 1

        return generated

    def _mark_modified(self):
        self.metadata["modified"] = datetime.now().isoformat()
    
    def to_dict(self):
        """Esporta per JSON"""
        return {
            "metadata": self.metadata,
            "points": [point.to_dict() for point in self.points.values()],
            "next_id": self.next_id
        }
    
    def from_dict(self, data):
        """Importa da JSON"""
        self.clear()

        if "metadata" in data:
            self.metadata.update(data["metadata"])

        if "points" in data:
            for point_data in data["points"]:
                point = Point.from_dict(point_data)
                self.points[point.id] = point

        self.next_id = data.get("next_id", 1)

        # Correggi ID se necessario
        if self.points:
            max_id = max(self.points.keys())
            if max_id >= self.next_id:
                self.next_id = max_id + 1

    def export_excel(self, filename):
        """Esporta coordinate in formato Excel (.xlsx)"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl non disponibile. Installare con: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Coordinate Fili Fissi"

        # Stili
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ["ID", "X (m)", "Y (m)", "Z (m)", "Livello", "Descrizione"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Dati punti
        for row_idx, point in enumerate(self.get_point_list(), start=2):
            ws.cell(row=row_idx, column=1, value=point.id).border = border
            ws.cell(row=row_idx, column=2, value=point.x).border = border
            ws.cell(row=row_idx, column=2, value=point.x).number_format = '0.0000'
            ws.cell(row=row_idx, column=3, value=point.y).border = border
            ws.cell(row=row_idx, column=3, value=point.y).number_format = '0.0000'
            ws.cell(row=row_idx, column=4, value=point.z).border = border
            ws.cell(row=row_idx, column=4, value=point.z).number_format = '0.000'
            ws.cell(row=row_idx, column=5, value=point.level).border = border
            ws.cell(row=row_idx, column=6, value=point.description).border = border

        # Larghezza colonne
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30

        # Foglio Metadata
        ws_meta = wb.create_sheet("Info Progetto")
        ws_meta.column_dimensions['A'].width = 20
        ws_meta.column_dimensions['B'].width = 50

        meta_data = [
            ("Progetto", self.metadata.get('project_name', '')),
            ("Località", self.metadata.get('location', '')),
            ("Autore", self.metadata.get('author', '')),
            ("Registrazione", self.metadata.get('registration', '')),
            ("Tipo", self.metadata.get('project_type', '')),
            ("Creato", self.metadata.get('created', '')),
            ("Modificato", self.metadata.get('modified', '')),
            ("Note", self.metadata.get('notes', ''))
        ]

        for row_idx, (key, value) in enumerate(meta_data, start=1):
            cell_key = ws_meta.cell(row=row_idx, column=1, value=key)
            cell_key.font = Font(bold=True)
            ws_meta.cell(row=row_idx, column=2, value=value)

        # Statistiche
        ws_stats = wb.create_sheet("Statistiche")
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 15

        bounds = self.get_bounds() if self.points else None
        if bounds:
            stats_data = [
                ("STATISTICHE COORDINATE", ""),
                ("Totale punti", len(self.points)),
                ("", ""),
                ("X min (m)", f"{bounds['x_min']:.4f}"),
                ("X max (m)", f"{bounds['x_max']:.4f}"),
                ("Range X (m)", f"{bounds['x_max'] - bounds['x_min']:.4f}"),
                ("", ""),
                ("Y min (m)", f"{bounds['y_min']:.4f}"),
                ("Y max (m)", f"{bounds['y_max']:.4f}"),
                ("Range Y (m)", f"{bounds['y_max'] - bounds['y_min']:.4f}"),
                ("", ""),
                ("Z min (m)", f"{bounds['z_min']:.3f}"),
                ("Z max (m)", f"{bounds['z_max']:.3f}"),
                ("Range Z (m)", f"{bounds['z_max'] - bounds['z_min']:.3f}"),
            ]

            for row_idx, (key, value) in enumerate(stats_data, start=1):
                cell_key = ws_stats.cell(row=row_idx, column=1, value=key)
                if key and not value:
                    cell_key.font = Font(bold=True, size=12)
                ws_stats.cell(row=row_idx, column=2, value=value)

        wb.save(filename)

    def import_excel(self, filename):
        """Importa coordinate da file Excel (.xlsx)"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl non disponibile. Installare con: pip install openpyxl")

        wb = load_workbook(filename, data_only=True)
        ws = wb.active

        imported = 0
        # Salta header (riga 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                # row = (ID, X, Y, Z, Livello, Descrizione)
                if row[1] is None or row[2] is None:  # X e Y obbligatori
                    continue

                x = float(row[1])
                y = float(row[2])
                z = float(row[3]) if row[3] is not None else 0.0
                level = str(row[4]) if row[4] else ""
                description = str(row[5]) if row[5] else ""

                self.add_point(x, y, z, description, level)
                imported += 1

            except (ValueError, TypeError, IndexError):
                continue

        wb.close()
        return imported

    def import_dxf(self, filename):
        """Importa coordinate da file DXF (AutoCAD)"""
        imported = 0

        if EZDXF_AVAILABLE:
            # Import avanzato con ezdxf
            try:
                doc = ezdxf.readfile(filename)
                msp = doc.modelspace()

                # Leggi POINT entities
                for entity in msp.query('POINT'):
                    location = entity.dxf.location
                    x, y, z = location.x, location.y, location.z
                    # Usa layer come livello se disponibile
                    level = entity.dxf.layer if hasattr(entity.dxf, 'layer') else ""
                    description = f"DXF Point"
                    self.add_point(x, y, z, description, level)
                    imported += 1

                # Leggi TEXT entities come descrizioni di punti
                text_entities = list(msp.query('TEXT'))
                for text_entity in text_entities:
                    insert = text_entity.dxf.insert
                    text_content = text_entity.dxf.text
                    x, y, z = insert.x, insert.y, insert.z
                    level = text_entity.dxf.layer if hasattr(text_entity.dxf, 'layer') else ""
                    # Cerca se esiste già un punto molto vicino
                    point_exists = False
                    for p in self.points.values():
                        dist = math.sqrt((p.x - x)**2 + (p.y - y)**2 + (p.z - z)**2)
                        if dist < 0.01:  # Stesso punto
                            # Aggiorna descrizione
                            if not p.description or p.description == "DXF Point":
                                p.description = text_content
                            point_exists = True
                            break

                    if not point_exists:
                        self.add_point(x, y, z, text_content, level)
                        imported += 1

                # Leggi INSERT entities (blocchi) come punti
                for entity in msp.query('INSERT'):
                    insert = entity.dxf.insert
                    x, y, z = insert.x, insert.y, insert.z
                    block_name = entity.dxf.name
                    level = entity.dxf.layer if hasattr(entity.dxf, 'layer') else ""
                    self.add_point(x, y, z, f"Block: {block_name}", level)
                    imported += 1

                return imported

            except Exception as e:
                # Fallback a parser manuale
                pass

        # Parser DXF manuale (fallback se ezdxf non disponibile o fallisce)
        with open(filename, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            # Cerca entità POINT
            if line == 'POINT':
                x, y, z = 0.0, 0.0, 0.0
                layer = ""
                # Leggi coordinate gruppo code 10, 20, 30
                j = i + 1
                while j < len(lines) and j < i + 50:  # Limita ricerca
                    code = lines[j].strip()
                    if j + 1 < len(lines):
                        value = lines[j + 1].strip()
                        if code == '8':  # Layer
                            layer = value
                        elif code == '10':  # X
                            try:
                                x = float(value)
                            except:
                                pass
                        elif code == '20':  # Y
                            try:
                                y = float(value)
                            except:
                                pass
                        elif code == '30':  # Z
                            try:
                                z = float(value)
                            except:
                                pass
                        elif code == '0':  # Nuova entità
                            break
                    j += 2

                self.add_point(x, y, z, "DXF Point", layer)
                imported += 1

            # Cerca entità TEXT
            elif line == 'TEXT':
                x, y, z = 0.0, 0.0, 0.0
                text_content = ""
                layer = ""
                j = i + 1
                while j < len(lines) and j < i + 50:
                    code = lines[j].strip()
                    if j + 1 < len(lines):
                        value = lines[j + 1].strip()
                        if code == '8':  # Layer
                            layer = value
                        elif code == '10':  # X
                            try:
                                x = float(value)
                            except:
                                pass
                        elif code == '20':  # Y
                            try:
                                y = float(value)
                            except:
                                pass
                        elif code == '30':  # Z
                            try:
                                z = float(value)
                            except:
                                pass
                        elif code == '1':  # Testo
                            text_content = value
                        elif code == '0':  # Nuova entità
                            break
                    j += 2

                if text_content:
                    self.add_point(x, y, z, text_content, layer)
                    imported += 1

            i += 1

        return imported

    def generate_pdf_report(self, filename):
        """Genera report PDF completo con coordinate e statistiche"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab non disponibile. Installare con: pip install reportlab")

        # Crea documento PDF
        doc = SimpleDocTemplate(filename, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)

        story = []
        styles = getSampleStyleSheet()

        # Stile custom per titolo
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#003366'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        # Stile custom per sottotitolo
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12
        )

        # === PAGINA 1: FRONTESPIZIO ===

        # Titolo
        story.append(Spacer(1, 3*cm))
        story.append(Paragraph("REPORT COORDINATE FILI FISSI", title_style))
        story.append(Spacer(1, 1*cm))

        # Informazioni progetto
        project_name = self.metadata.get('project_name', 'Non specificato')
        location = self.metadata.get('location', 'Non specificato')
        author = self.metadata.get('author', 'N/A')
        registration = self.metadata.get('registration', 'N/A')

        story.append(Paragraph(f"<b>Progetto:</b> {project_name}", styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"<b>Località:</b> {location}", styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"<b>Architetto:</b> {author}", styles['Normal']))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f"<b>Registrazione:</b> {registration}", styles['Normal']))
        story.append(Spacer(1, 1*cm))

        # Data creazione e modifica
        created = self.metadata.get('created', 'N/A')[:19]
        modified = self.metadata.get('modified', 'N/A')[:19]
        story.append(Paragraph(f"<b>Creato:</b> {created}", styles['Normal']))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f"<b>Modificato:</b> {modified}", styles['Normal']))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f"<b>Generato:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))

        # Note
        notes = self.metadata.get('notes', '')
        if notes:
            story.append(Spacer(1, 1*cm))
            story.append(Paragraph("<b>Note:</b>", styles['Normal']))
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph(notes, styles['Normal']))

        story.append(PageBreak())

        # === PAGINA 2: STATISTICHE ===

        story.append(Paragraph("STATISTICHE COORDINATE", subtitle_style))
        story.append(Spacer(1, 0.5*cm))

        # Tabella statistiche
        stats_data = [
            ["Descrizione", "Valore"],
            ["Totale punti coordinati", str(len(self.points))]
        ]

        if self.points:
            bounds = self.get_bounds()
            x_range = bounds['x_max'] - bounds['x_min']
            y_range = bounds['y_max'] - bounds['y_min']
            z_range = bounds['z_max'] - bounds['z_min']

            stats_data.extend([
                ["", ""],
                ["ESTENSIONE X", ""],
                [f"  X minima", f"{bounds['x_min']:.4f} m"],
                [f"  X massima", f"{bounds['x_max']:.4f} m"],
                [f"  Range X", f"{x_range:.4f} m"],
                ["", ""],
                ["ESTENSIONE Y", ""],
                [f"  Y minima", f"{bounds['y_min']:.4f} m"],
                [f"  Y massima", f"{bounds['y_max']:.4f} m"],
                [f"  Range Y", f"{y_range:.4f} m"],
                ["", ""],
                ["ESTENSIONE Z", ""],
                [f"  Z minima", f"{bounds['z_min']:.3f} m"],
                [f"  Z massima", f"{bounds['z_max']:.3f} m"],
                [f"  Range Z", f"{z_range:.3f} m"],
            ])

        stats_table = Table(stats_data, colWidths=[10*cm, 6*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ]))

        story.append(stats_table)
        story.append(Spacer(1, 1*cm))

        # Livelli
        levels = self.get_levels()
        if levels:
            story.append(Paragraph("LIVELLI DEFINITI", subtitle_style))
            story.append(Spacer(1, 0.5*cm))

            levels_data = [["Livello", "Numero Punti"]]
            for level in levels:
                count = len(self.get_points_by_level(level))
                levels_data.append([level, str(count)])

            levels_table = Table(levels_data, colWidths=[10*cm, 6*cm])
            levels_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            story.append(levels_table)

        story.append(PageBreak())

        # === PAGINA 3+: TABELLA COORDINATE ===

        story.append(Paragraph("ELENCO COORDINATE FILI FISSI", subtitle_style))
        story.append(Spacer(1, 0.5*cm))

        # Tabella coordinate
        coord_data = [["ID", "X (m)", "Y (m)", "Z (m)", "Livello", "Descrizione"]]

        for point in self.get_point_list():
            coord_data.append([
                str(point.id),
                f"{point.x:.4f}",
                f"{point.y:.4f}",
                f"{point.z:.3f}",
                point.level,
                point.description[:30]  # Limita lunghezza
            ])

        # Larghezze colonne
        col_widths = [1.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm, 6*cm]

        coord_table = Table(coord_data, colWidths=col_widths, repeatRows=1)
        coord_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))

        story.append(coord_table)

        # Genera PDF
        doc.build(story)


class MetadataDialog(QDialog):
    """Dialog per modifica metadata progetto"""

    def __init__(self, metadata, parent=None):
        super().__init__(parent)
        self.metadata = metadata.copy()
        self.setWindowTitle("Informazioni Progetto")
        self.setup_ui()

    def setup_ui(self):
        layout = QGridLayout()

        # Nome progetto
        layout.addWidget(QLabel("Nome Progetto:"), 0, 0)
        self.project_name = QLineEdit()
        self.project_name.setText(self.metadata.get('project_name', ''))
        layout.addWidget(self.project_name, 0, 1)

        # Location
        layout.addWidget(QLabel("Localit\u00e0:"), 1, 0)
        self.location = QLineEdit()
        self.location.setText(self.metadata.get('location', ''))
        layout.addWidget(self.location, 1, 1)

        # Note
        layout.addWidget(QLabel("Note:"), 2, 0)
        self.notes = QTextEdit()
        self.notes.setPlainText(self.metadata.get('notes', ''))
        self.notes.setMaximumHeight(100)
        layout.addWidget(self.notes, 2, 1)

        # Autore (read-only)
        layout.addWidget(QLabel("Autore:"), 3, 0)
        author_label = QLabel(self.metadata.get('author', ''))
        author_label.setStyleSheet("color: gray;")
        layout.addWidget(author_label, 3, 1)

        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons, 4, 0, 1, 2)

        self.setLayout(layout)
        self.setMinimumWidth(500)

    def get_metadata(self):
        """Restituisce metadata aggiornato"""
        self.metadata['project_name'] = self.project_name.text().strip()
        self.metadata['location'] = self.location.text().strip()
        self.metadata['notes'] = self.notes.toPlainText().strip()
        return self.metadata


class Annotation:
    """Classe base per annotazioni grafiche"""
    def __init__(self, annotation_type, color='red', linewidth=2):
        self.type = annotation_type
        self.color = color
        self.linewidth = linewidth

    def to_dict(self):
        return {
            'type': self.type,
            'color': self.color,
            'linewidth': self.linewidth
        }


class LineAnnotation(Annotation):
    """Annotazione linea tra due punti"""
    def __init__(self, p1_id, p2_id, color='red', linewidth=2, style='solid'):
        super().__init__('line', color, linewidth)
        self.p1_id = p1_id
        self.p2_id = p2_id
        self.style = style  # 'solid', 'dashed', 'dotted'

    def to_dict(self):
        data = super().to_dict()
        data.update({
            'p1_id': self.p1_id,
            'p2_id': self.p2_id,
            'style': self.style
        })
        return data

    @classmethod
    def from_dict(cls, data):
        return cls(data['p1_id'], data['p2_id'],
                  data.get('color', 'red'),
                  data.get('linewidth', 2),
                  data.get('style', 'solid'))


class DimensionAnnotation(Annotation):
    """Annotazione quota/dimensione tra due punti"""
    def __init__(self, p1_id, p2_id, offset=0.5, color='blue', show_text=True):
        super().__init__('dimension', color, 1)
        self.p1_id = p1_id
        self.p2_id = p2_id
        self.offset = offset  # Offset della linea di quota
        self.show_text = show_text

    def to_dict(self):
        data = super().to_dict()
        data.update({
            'p1_id': self.p1_id,
            'p2_id': self.p2_id,
            'offset': self.offset,
            'show_text': self.show_text
        })
        return data

    @classmethod
    def from_dict(cls, data):
        return cls(data['p1_id'], data['p2_id'],
                  data.get('offset', 0.5),
                  data.get('color', 'blue'),
                  data.get('show_text', True))


class PlotDialog(QDialog):
    """Dialog per visualizzazione grafica 2D/3D dei punti"""

    def __init__(self, model, parent=None):
        super().__init__(parent)
        self.model = model
        self.setWindowTitle("Vista Grafica Coordinate")
        self.setMinimumSize(900, 700)
        self.annotations = []  # Lista annotazioni
        self.setup_ui()

    def setup_ui(self):
        if not MATPLOTLIB_AVAILABLE:
            layout = QVBoxLayout()
            label = QLabel("Matplotlib non disponibile.\nInstallare con: pip install matplotlib")
            label.setAlignment(Qt.AlignCenter)
            layout.addWidget(label)
            close_btn = QPushButton("Chiudi")
            close_btn.clicked.connect(self.close)
            layout.addWidget(close_btn)
            self.setLayout(layout)
            return

        layout = QVBoxLayout()

        # Toolbar controlli
        controls = QHBoxLayout()

        controls.addWidget(QLabel("Livello:"))
        self.level_combo = QComboBox()
        self.level_combo.addItem("Tutti")
        for level in self.model.get_levels():
            self.level_combo.addItem(level)
        self.level_combo.currentTextChanged.connect(self.update_plot)
        controls.addWidget(self.level_combo)

        controls.addWidget(QLabel("Vista:"))
        self.view_combo = QComboBox()
        self.view_combo.addItems(["2D Planimetria (XY)", "Vista XZ", "Vista YZ", "3D"])
        self.view_combo.currentTextChanged.connect(self.update_plot)
        controls.addWidget(self.view_combo)

        self.show_labels = QCheckBox("Mostra Etichette")
        self.show_labels.setChecked(True)
        self.show_labels.stateChanged.connect(self.update_plot)
        controls.addWidget(self.show_labels)

        self.show_grid = QCheckBox("Griglia")
        self.show_grid.setChecked(True)
        self.show_grid.stateChanged.connect(self.update_plot)
        controls.addWidget(self.show_grid)

        export_btn = QPushButton("Esporta PNG")
        export_btn.clicked.connect(self.export_image)
        controls.addWidget(export_btn)

        controls.addStretch()
        layout.addLayout(controls)

        # Toolbar annotazioni
        annot_layout = QHBoxLayout()
        annot_layout.addWidget(QLabel("Annotazioni:"))

        add_line_btn = QPushButton("Aggiungi Linea")
        add_line_btn.clicked.connect(self.add_line_annotation)
        annot_layout.addWidget(add_line_btn)

        add_dim_btn = QPushButton("Aggiungi Quota")
        add_dim_btn.clicked.connect(self.add_dimension_annotation)
        annot_layout.addWidget(add_dim_btn)

        clear_annot_btn = QPushButton("Pulisci Annotazioni")
        clear_annot_btn.clicked.connect(self.clear_annotations)
        annot_layout.addWidget(clear_annot_btn)

        annot_layout.addStretch()
        layout.addLayout(annot_layout)

        # Canvas matplotlib
        self.figure = Figure(figsize=(10, 8))
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)

        # Toolbar navigazione
        self.toolbar = NavigationToolbar(self.canvas, self)
        layout.addWidget(self.toolbar)

        # Bottone chiudi
        close_btn = QPushButton("Chiudi")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)

        self.setLayout(layout)
        self.update_plot()

    def update_plot(self):
        """Aggiorna il grafico"""
        if not MATPLOTLIB_AVAILABLE:
            return

        self.figure.clear()

        # Filtra punti per livello
        filter_level = self.level_combo.currentText()
        if filter_level == "Tutti":
            points = list(self.model.points.values())
        else:
            points = self.model.get_points_by_level(filter_level)

        if not points:
            ax = self.figure.add_subplot(111)
            ax.text(0.5, 0.5, 'Nessun punto da visualizzare',
                   ha='center', va='center', fontsize=14)
            self.canvas.draw()
            return

        view_type = self.view_combo.currentText()

        if "3D" in view_type:
            ax = self.figure.add_subplot(111, projection='3d')
            self.plot_3d(ax, points)
        else:
            ax = self.figure.add_subplot(111)
            self.plot_2d(ax, points, view_type)

        self.canvas.draw()

    def plot_2d(self, ax, points, view_type):
        """Plot 2D"""
        # Colori per livelli
        levels = list(set(p.level for p in points if p.level))
        colors = plt.cm.tab10(range(len(levels))) if levels else ['blue']
        level_colors = {level: colors[i % len(colors)] for i, level in enumerate(levels)}

        # Seleziona assi
        if "XY" in view_type:
            x_vals = [p.x for p in points]
            y_vals = [p.y for p in points]
            xlabel, ylabel = 'X [m]', 'Y [m]'
        elif "XZ" in view_type:
            x_vals = [p.x for p in points]
            y_vals = [p.z for p in points]
            xlabel, ylabel = 'X [m]', 'Z [m]'
        else:  # YZ
            x_vals = [p.y for p in points]
            y_vals = [p.z for p in points]
            xlabel, ylabel = 'Y [m]', 'Z [m]'

        # Plot punti colorati per livello
        for point in points:
            if "XY" in view_type:
                x, y = point.x, point.y
            elif "XZ" in view_type:
                x, y = point.x, point.z
            else:
                x, y = point.y, point.z

            color = level_colors.get(point.level, 'blue') if point.level else 'blue'
            ax.scatter(x, y, c=[color], s=100, alpha=0.6, edgecolors='black', linewidths=1)

            # Etichette
            if self.show_labels.isChecked():
                label = point.description if point.description else f"P{point.id}"
                ax.annotate(label, (x, y), xytext=(5, 5),
                           textcoords='offset points', fontsize=8)

        ax.set_xlabel(xlabel, fontsize=12)
        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_title(f'Vista {view_type} - {len(points)} punti', fontsize=14, fontweight='bold')
        ax.grid(self.show_grid.isChecked(), alpha=0.3)
        ax.axis('equal')

        # Legenda livelli
        if levels:
            from matplotlib.patches import Patch
            legend_elements = [Patch(facecolor=level_colors[level], label=level)
                             for level in levels]
            ax.legend(handles=legend_elements, loc='best', fontsize=9)

        # Renderizza annotazioni
        self.render_annotations_2d(ax, view_type)

    def render_annotations_2d(self, ax, view_type):
        """Renderizza annotazioni grafiche nel plot 2D"""
        for annot in self.annotations:
            if isinstance(annot, LineAnnotation):
                # Disegna linea
                if annot.p1_id in self.model.points and annot.p2_id in self.model.points:
                    p1 = self.model.points[annot.p1_id]
                    p2 = self.model.points[annot.p2_id]

                    # Seleziona coordinate corrette per vista
                    if "XY" in view_type:
                        x1, y1 = p1.x, p1.y
                        x2, y2 = p2.x, p2.y
                    elif "XZ" in view_type:
                        x1, y1 = p1.x, p1.z
                        x2, y2 = p2.x, p2.z
                    else:  # YZ
                        x1, y1 = p1.y, p1.z
                        x2, y2 = p2.y, p2.z

                    linestyle = '--' if annot.style == 'dashed' else (':' if annot.style == 'dotted' else '-')
                    ax.plot([x1, x2], [y1, y2], color=annot.color,
                           linewidth=annot.linewidth, linestyle=linestyle, alpha=0.7)

            elif isinstance(annot, DimensionAnnotation):
                # Disegna quota
                if annot.p1_id in self.model.points and annot.p2_id in self.model.points:
                    p1 = self.model.points[annot.p1_id]
                    p2 = self.model.points[annot.p2_id]

                    # Coordinate per vista
                    if "XY" in view_type:
                        x1, y1 = p1.x, p1.y
                        x2, y2 = p2.x, p2.y
                    elif "XZ" in view_type:
                        x1, y1 = p1.x, p1.z
                        x2, y2 = p2.x, p2.z
                    else:  # YZ
                        x1, y1 = p1.y, p1.z
                        x2, y2 = p2.y, p2.z

                    # Calcola direzione perpendicolare per offset
                    dx = x2 - x1
                    dy = y2 - y1
                    length = math.sqrt(dx**2 + dy**2)

                    if length > 0:
                        # Vettore perpendicolare normalizzato
                        perp_x = -dy / length * annot.offset
                        perp_y = dx / length * annot.offset

                        # Linea di quota offset
                        ox1, oy1 = x1 + perp_x, y1 + perp_y
                        ox2, oy2 = x2 + perp_x, y2 + perp_y

                        # Disegna linea di quota
                        ax.plot([ox1, ox2], [oy1, oy2], color=annot.color,
                               linewidth=1, linestyle='-', alpha=0.7)

                        # Linee di estensione
                        ax.plot([x1, ox1], [y1, oy1], color=annot.color,
                               linewidth=0.5, linestyle=':', alpha=0.5)
                        ax.plot([x2, ox2], [y2, oy2], color=annot.color,
                               linewidth=0.5, linestyle=':', alpha=0.5)

                        # Frecce alle estremità
                        ax.annotate('', xy=(ox1, oy1), xytext=(ox1 + dx * 0.05, oy1 + dy * 0.05),
                                   arrowprops=dict(arrowstyle='<-', color=annot.color, lw=1))
                        ax.annotate('', xy=(ox2, oy2), xytext=(ox2 - dx * 0.05, oy2 - dy * 0.05),
                                   arrowprops=dict(arrowstyle='<-', color=annot.color, lw=1))

                        # Testo con dimensione
                        if annot.show_text:
                            mid_x = (ox1 + ox2) / 2
                            mid_y = (oy1 + oy2) / 2
                            ax.text(mid_x, mid_y, f'{length:.3f}m',
                                   fontsize=9, color=annot.color, fontweight='bold',
                                   ha='center', va='bottom',
                                   bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.8))

    def plot_3d(self, ax, points):
        """Plot 3D"""
        levels = list(set(p.level for p in points if p.level))
        colors = plt.cm.tab10(range(len(levels))) if levels else ['blue']
        level_colors = {level: colors[i % len(colors)] for i, level in enumerate(levels)}

        for point in points:
            color = level_colors.get(point.level, 'blue') if point.level else 'blue'
            ax.scatter(point.x, point.y, point.z, c=[color], s=100,
                      alpha=0.6, edgecolors='black', linewidths=1)

            if self.show_labels.isChecked():
                label = point.description if point.description else f"P{point.id}"
                ax.text(point.x, point.y, point.z, label, fontsize=8)

        ax.set_xlabel('X [m]', fontsize=12)
        ax.set_ylabel('Y [m]', fontsize=12)
        ax.set_zlabel('Z [m]', fontsize=12)
        ax.set_title(f'Vista 3D - {len(points)} punti', fontsize=14, fontweight='bold')
        ax.grid(self.show_grid.isChecked(), alpha=0.3)

        # Legenda
        if levels:
            from matplotlib.patches import Patch
            legend_elements = [Patch(facecolor=level_colors[level], label=level)
                             for level in levels]
            ax.legend(handles=legend_elements, loc='best', fontsize=9)

    def export_image(self):
        """Esporta grafico come PNG"""
        if not MATPLOTLIB_AVAILABLE:
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Esporta Immagine", "coordinate_plot.png",
            "Immagine PNG (*.png);;Tutti i file (*)"
        )

        if filename:
            try:
                self.figure.savefig(filename, dpi=300, bbox_inches='tight')
                QMessageBox.information(self, "Esportazione Completata",
                                      f"Immagine esportata in:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Esportazione",
                                   f"Impossibile esportare immagine:\n{e}")

    def add_line_annotation(self):
        """Aggiungi annotazione linea tra due punti"""
        if len(self.model.points) < 2:
            QMessageBox.warning(self, "Attenzione", "Servono almeno 2 punti per creare una linea")
            return

        # Lista ID punti
        point_ids = [str(p.id) for p in self.model.get_point_list()]

        # Selezione punto 1
        id1_str, ok1 = QInputDialog.getItem(
            self, "Linea - Punto 1", "Seleziona primo punto (ID):", point_ids, 0, False
        )
        if not ok1:
            return
        id1 = int(id1_str)

        # Selezione punto 2
        id2_str, ok2 = QInputDialog.getItem(
            self, "Linea - Punto 2", "Seleziona secondo punto (ID):", point_ids, 0, False
        )
        if not ok2:
            return
        id2 = int(id2_str)

        if id1 == id2:
            QMessageBox.warning(self, "Attenzione", "Seleziona due punti diversi")
            return

        # Stile linea
        styles = ["Continua (solid)", "Tratteggiata (dashed)", "Punteggiata (dotted)"]
        style_sel, ok3 = QInputDialog.getItem(
            self, "Stile Linea", "Seleziona stile:", styles, 0, False
        )
        if not ok3:
            return

        style_map = {"Continua (solid)": "solid", "Tratteggiata (dashed)": "dashed", "Punteggiata (dotted)": "dotted"}
        style = style_map[style_sel]

        # Crea annotazione
        annot = LineAnnotation(id1, id2, color='red', linewidth=2, style=style)
        self.annotations.append(annot)
        self.update_plot()

        QMessageBox.information(self, "Linea Aggiunta",
                              f"Linea aggiunta tra punto {id1} e punto {id2}")

    def add_dimension_annotation(self):
        """Aggiungi annotazione quota tra due punti"""
        if len(self.model.points) < 2:
            QMessageBox.warning(self, "Attenzione", "Servono almeno 2 punti per creare una quota")
            return

        # Lista ID punti
        point_ids = [str(p.id) for p in self.model.get_point_list()]

        # Selezione punto 1
        id1_str, ok1 = QInputDialog.getItem(
            self, "Quota - Punto 1", "Seleziona primo punto (ID):", point_ids, 0, False
        )
        if not ok1:
            return
        id1 = int(id1_str)

        # Selezione punto 2
        id2_str, ok2 = QInputDialog.getItem(
            self, "Quota - Punto 2", "Seleziona secondo punto (ID):", point_ids, 0, False
        )
        if not ok2:
            return
        id2 = int(id2_str)

        if id1 == id2:
            QMessageBox.warning(self, "Attenzione", "Seleziona due punti diversi")
            return

        # Offset quota
        offset, ok3 = QInputDialog.getDouble(
            self, "Offset Quota", "Offset linea di quota (m):", 0.5, 0.1, 10.0, 2
        )
        if not ok3:
            return

        # Crea annotazione
        annot = DimensionAnnotation(id1, id2, offset=offset, color='blue', show_text=True)
        self.annotations.append(annot)
        self.update_plot()

        QMessageBox.information(self, "Quota Aggiunta",
                              f"Quota aggiunta tra punto {id1} e punto {id2}")

    def clear_annotations(self):
        """Pulisci tutte le annotazioni"""
        if not self.annotations:
            QMessageBox.information(self, "Annotazioni", "Nessuna annotazione da rimuovere")
            return

        reply = QMessageBox.question(
            self, 'Conferma',
            f'Rimuovere tutte le {len(self.annotations)} annotazioni?',
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.annotations.clear()
            self.update_plot()
            QMessageBox.information(self, "Annotazioni Rimosse",
                                  "Tutte le annotazioni sono state rimosse")


class CoordinateInputPanel(QGroupBox):
    """Pannello input coordinate semplificato"""
    
    def __init__(self, model, parent=None):
        super().__init__("Input Coordinate", parent)
        self.model = model
        self.setup_ui()
    
    def setup_ui(self):
        layout = QGridLayout()
        
        # Coordinate principali
        layout.addWidget(QLabel("X (m):"), 0, 0)
        self.x_spin = QDoubleSpinBox()
        self.x_spin.setRange(-10000, 10000)
        self.x_spin.setDecimals(4)
        self.x_spin.setSingleStep(0.1)
        self.x_spin.setMinimumWidth(120)
        layout.addWidget(self.x_spin, 0, 1)
        
        layout.addWidget(QLabel("Y (m):"), 0, 2)
        self.y_spin = QDoubleSpinBox()
        self.y_spin.setRange(-10000, 10000)
        self.y_spin.setDecimals(4)
        self.y_spin.setSingleStep(0.1)
        self.y_spin.setMinimumWidth(120)
        layout.addWidget(self.y_spin, 0, 3)
        
        layout.addWidget(QLabel("Z (m):"), 1, 0)
        self.z_spin = QDoubleSpinBox()
        self.z_spin.setRange(-1000, 1000)
        self.z_spin.setDecimals(3)
        self.z_spin.setSingleStep(0.1)
        self.z_spin.setMinimumWidth(120)
        layout.addWidget(self.z_spin, 1, 1)
        
        # Descrizione opzionale
        layout.addWidget(QLabel("Descrizione:"), 1, 2)
        self.desc_edit = QLineEdit()
        self.desc_edit.setPlaceholderText("Opzionale - es: Pilastro A1")
        layout.addWidget(self.desc_edit, 1, 3)
        
        # Pulsanti
        button_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("Aggiungi Punto")
        self.add_btn.setDefault(True)
        self.add_btn.clicked.connect(self.add_point)
        button_layout.addWidget(self.add_btn)
        
        self.clear_btn = QPushButton("Reset")
        self.clear_btn.clicked.connect(self.clear_fields)
        button_layout.addWidget(self.clear_btn)
        
        layout.addLayout(button_layout, 2, 0, 1, 4)
        self.setLayout(layout)
        
        # Focus su X per input rapido
        self.x_spin.setFocus()
        
        # Enter per aggiungere (solo LineEdit supporta returnPressed)
        self.desc_edit.returnPressed.connect(self.add_point)
    
    def add_point(self):
        """Aggiunge punto alle coordinate"""
        x = self.x_spin.value()
        y = self.y_spin.value()
        z = self.z_spin.value()
        description = self.desc_edit.text().strip()
        
        point_id = self.model.add_point(x, y, z, description)
        
        # Incrementa automaticamente X per input sequenziale
        self.x_spin.setValue(x + 1.0)
        self.desc_edit.clear()
        self.x_spin.setFocus()
        
        return point_id
    
    def clear_fields(self):
        """Reset tutti i campi"""
        self.x_spin.setValue(0.0)
        self.y_spin.setValue(0.0)
        self.z_spin.setValue(0.0)
        self.desc_edit.clear()
        self.x_spin.setFocus()
    
    def keyPressEvent(self, event):
        """Enter = aggiungi punto"""
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            self.add_point()
        super().keyPressEvent(event)


class CoordinateTable(QTableWidget):
    """Tabella coordinate con editing inline"""
    
    def __init__(self, model, parent=None):
        super().__init__(parent)
        self.model = model
        self.setup_table()
    
    def setup_table(self):
        self.setColumnCount(6)
        self.setHorizontalHeaderLabels(['ID', 'X (m)', 'Y (m)', 'Z (m)', 'Livello', 'Descrizione'])

        # Configurazione header
        header = self.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID fisso
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # X
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Y
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Z
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Livello
        header.setStretchLastSection(True)  # Descrizione espandibile

        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(True)

        # Connetti modifiche
        self.cellChanged.connect(self.on_cell_changed)
    
    def refresh(self, filter_level=None):
        """Aggiorna tabella completa con filtro opzionale per livello"""
        points = self.model.get_point_list()

        # Applica filtro livello se specificato
        if filter_level and filter_level != "Tutti":
            points = [p for p in points if p.level == filter_level]

        self.setRowCount(len(points))

        # Blocca segnali durante refresh
        self.blockSignals(True)

        for row, point in enumerate(points):
            # ID (solo lettura)
            id_item = QTableWidgetItem(str(point.id))
            id_item.setFlags(id_item.flags() & ~Qt.ItemIsEditable)
            self.setItem(row, 0, id_item)

            # Coordinate (modificabili con precisione)
            self.setItem(row, 1, QTableWidgetItem(f"{point.x:.4f}"))
            self.setItem(row, 2, QTableWidgetItem(f"{point.y:.4f}"))
            self.setItem(row, 3, QTableWidgetItem(f"{point.z:.3f}"))

            # Livello
            self.setItem(row, 4, QTableWidgetItem(point.level))

            # Descrizione
            self.setItem(row, 5, QTableWidgetItem(point.description))

        self.blockSignals(False)
    
    def on_cell_changed(self, row, col):
        """Gestisce modifiche dirette in tabella"""
        id_item = self.item(row, 0)
        if not id_item:
            return

        point_id = int(id_item.text())
        item = self.item(row, col)

        try:
            if col == 1:  # X
                self.model.update_point(point_id, x=float(item.text()))
            elif col == 2:  # Y
                self.model.update_point(point_id, y=float(item.text()))
            elif col == 3:  # Z
                self.model.update_point(point_id, z=float(item.text()))
            elif col == 4:  # Livello
                self.model.update_point(point_id, level=item.text())
            elif col == 5:  # Descrizione
                self.model.update_point(point_id, description=item.text())
        except ValueError as e:
            # Valore non valido, ripristina
            QMessageBox.warning(self, "Errore", f"Valore non valido: {e}")
            self.refresh()
    
    def keyPressEvent(self, event):
        """Elimina punto con DEL"""
        if event.key() == Qt.Key_Delete:
            current_row = self.currentRow()
            if current_row >= 0:
                id_item = self.item(current_row, 0)
                if id_item:
                    point_id = int(id_item.text())
                    reply = QMessageBox.question(
                        self, 'Conferma Eliminazione', 
                        f'Eliminare il punto ID {point_id}?',
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if reply == QMessageBox.Yes:
                        self.model.delete_point(point_id)
                        self.refresh()
        super().keyPressEvent(event)


class MainWindow(QMainWindow):
    """Finestra principale per input coordinate fili fissi"""
    
    def __init__(self):
        super().__init__()
        self.model = CoordinateModel()
        self.setup_ui()
        self.setup_menus()
        self.setWindowTitle("Input Coordinate Fili Fissi - Arch. Michelangelo Bartolotta")
    
    def setup_ui(self):
        self.setMinimumSize(1000, 600)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # Pannello input semplificato
        self.input_panel = CoordinateInputPanel(self.model)
        self.input_panel.add_btn.clicked.connect(self.refresh_display)
        layout.addWidget(self.input_panel)
        
        # Tabella coordinate
        table_group = QGroupBox("Elenco Coordinate")
        table_layout = QVBoxLayout()
        
        self.table = CoordinateTable(self.model)
        table_layout.addWidget(self.table)
        
        table_group.setLayout(table_layout)
        layout.addWidget(table_group)
        
        # Barra informazioni e filtri
        info_layout = QHBoxLayout()

        self.info_label = QLabel("Punti: 0")
        self.info_label.setFont(QFont("Arial", 10, QFont.Bold))
        info_layout.addWidget(self.info_label)

        info_layout.addStretch()

        # Filtro livello
        info_layout.addWidget(QLabel("Filtro Livello:"))
        self.level_filter = QComboBox()
        self.level_filter.addItem("Tutti")
        self.level_filter.currentTextChanged.connect(self.on_level_filter_changed)
        self.level_filter.setMinimumWidth(120)
        info_layout.addWidget(self.level_filter)

        info_layout.addStretch()

        self.bounds_label = QLabel("")
        self.bounds_label.setFont(QFont("Arial", 9))
        info_layout.addWidget(self.bounds_label)

        layout.addLayout(info_layout)

        self.refresh_display()
    
    def setup_menus(self):
        menubar = self.menuBar()

        # Menu File
        file_menu = menubar.addMenu('File')
        file_menu.addAction('Nuovo Progetto', self.new_project, 'Ctrl+N')
        file_menu.addSeparator()
        file_menu.addAction('Apri...', self.open_file, 'Ctrl+O')
        file_menu.addAction('Salva...', self.save_file, 'Ctrl+S')
        file_menu.addSeparator()
        file_menu.addAction('Importa CSV...', self.import_csv)
        file_menu.addAction('Esporta CSV...', self.export_csv)
        file_menu.addSeparator()
        file_menu.addAction('Importa Excel...', self.import_excel)
        file_menu.addAction('Esporta Excel...', self.export_excel)
        file_menu.addSeparator()
        file_menu.addAction('Importa DXF...', self.import_dxf_ui)
        file_menu.addAction('Esporta DXF...', self.export_dxf)
        file_menu.addSeparator()
        file_menu.addAction('Genera Report PDF...', self.generate_pdf_report_ui, 'Ctrl+P')
        file_menu.addSeparator()
        file_menu.addAction('Informazioni Progetto...', self.edit_metadata, 'Ctrl+I')
        file_menu.addSeparator()
        file_menu.addAction('Esci', self.close, 'Ctrl+Q')

        # Menu Modifica
        edit_menu = menubar.addMenu('Modifica')
        self.undo_action = edit_menu.addAction('Annulla', self.undo_last_action, 'Ctrl+Z')
        self.redo_action = edit_menu.addAction('Ripristina', self.redo_last_action, 'Ctrl+Y')
        edit_menu.addSeparator()
        edit_menu.addAction('Duplica Punto Selezionato', self.duplicate_selected_point, 'Ctrl+D')
        edit_menu.addAction('Trasla Tutte le Coordinate...', self.translate_coordinates)
        edit_menu.addSeparator()
        edit_menu.addAction('Specchia Punti...', self.mirror_points_dialog)
        edit_menu.addAction('Snap to Grid...', self.snap_to_grid_dialog)
        edit_menu.addSeparator()
        edit_menu.addAction('Trova Punti Duplicati', self.find_duplicate_points)

        # Menu Livelli
        levels_menu = menubar.addMenu('Livelli')
        levels_menu.addAction('Assegna Livello a Selezione...', self.assign_level_dialog)
        levels_menu.addAction('Copia Livello...', self.copy_level_dialog)
        levels_menu.addSeparator()
        levels_menu.addAction('Gestione Livelli...', self.manage_levels_dialog)

        # Menu Genera
        generate_menu = menubar.addMenu('Genera')
        generate_menu.addAction('Griglia Regolare...', self.generate_grid_dialog)
        generate_menu.addAction('Interpiano Automatico...', self.generate_floors_dialog)

        # Menu Selezione
        selection_menu = menubar.addMenu('Selezione')
        selection_menu.addAction('Seleziona per Area...', self.select_by_area_dialog)
        selection_menu.addAction('Seleziona per Quota Z...', self.select_by_z_dialog)
        selection_menu.addSeparator()
        selection_menu.addAction('Trasla Selezione...', self.translate_selection_dialog)
        selection_menu.addAction('Elimina Selezione', self.delete_selection)

        # Menu Visualizza
        view_menu = menubar.addMenu('Visualizza')
        view_menu.addAction('Vista Grafica 2D/3D...', self.show_plot_window, 'Ctrl+G')

        # Menu Strumenti
        tools_menu = menubar.addMenu('Strumenti')
        tools_menu.addAction('Calcola Distanza tra Punti', self.calculate_distance_dialog)
        tools_menu.addSeparator()
        tools_menu.addAction('Statistiche Progetto', self.show_statistics)
        tools_menu.addAction('Info Coordinate', self.show_coordinate_info)
        tools_menu.addSeparator()
        tools_menu.addAction('Elimina Tutto', self.clear_all_points)

        # Menu Aiuto
        help_menu = menubar.addMenu('Aiuto')
        help_menu.addAction('Informazioni', self.show_about)
    
    def refresh_display(self):
        """Aggiorna tutte le visualizzazioni"""
        current_filter = self.level_filter.currentText()
        self.table.refresh(current_filter if current_filter != "Tutti" else None)
        count = len(self.model.points)
        self.info_label.setText(f"Punti: {count}")

        # Aggiorna combo livelli
        self.level_filter.blockSignals(True)
        current_selection = self.level_filter.currentText()
        self.level_filter.clear()
        self.level_filter.addItem("Tutti")
        for level in self.model.get_levels():
            self.level_filter.addItem(level)
        # Ripristina selezione se esiste ancora
        index = self.level_filter.findText(current_selection)
        if index >= 0:
            self.level_filter.setCurrentIndex(index)
        self.level_filter.blockSignals(False)

        # Aggiorna bounds
        if count > 0:
            bounds = self.model.get_bounds()
            bounds_text = (f"Range: X({bounds['x_min']:.2f}÷{bounds['x_max']:.2f}) "
                          f"Y({bounds['y_min']:.2f}÷{bounds['y_max']:.2f}) "
                          f"Z({bounds['z_min']:.2f}÷{bounds['z_max']:.2f})")
            self.bounds_label.setText(bounds_text)
        else:
            self.bounds_label.setText("")

        # Aggiorna stato Undo/Redo
        self.update_undo_redo_actions()

    def update_undo_redo_actions(self):
        """Aggiorna abilitazione azioni Undo/Redo"""
        self.undo_action.setEnabled(self.model.can_undo())
        self.redo_action.setEnabled(self.model.can_redo())

        # Aggiorna testo con numero operazioni
        undo_count = len(self.model.undo_stack)
        redo_count = len(self.model.redo_stack)
        self.undo_action.setText(f"Annulla ({undo_count})" if undo_count > 0 else "Annulla")
        self.redo_action.setText(f"Ripristina ({redo_count})" if redo_count > 0 else "Ripristina")

    def undo_last_action(self):
        """Annulla l'ultima azione"""
        if self.model.undo():
            self.refresh_display()
            QMessageBox.information(self, "Annulla",
                                  f"Operazione annullata. Rimaste {len(self.model.undo_stack)} operazioni.")
        else:
            QMessageBox.information(self, "Annulla", "Nessuna operazione da annullare")

    def redo_last_action(self):
        """Ripristina l'ultima azione annullata"""
        if self.model.redo():
            self.refresh_display()
            QMessageBox.information(self, "Ripristina",
                                  f"Operazione ripristinata. Disponibili {len(self.model.redo_stack)} ripristini.")
        else:
            QMessageBox.information(self, "Ripristina", "Nessuna operazione da ripristinare")

    def on_level_filter_changed(self, level):
        """Gestisce cambio filtro livello"""
        self.table.refresh(level if level != "Tutti" else None)
    
    def new_project(self):
        """Nuovo progetto"""
        if len(self.model.points) > 0:
            reply = QMessageBox.question(
                self, 'Nuovo Progetto', 
                'Eliminare tutte le coordinate esistenti?',
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.No:
                return
        
        self.model.clear()
        self.refresh_display()
        QMessageBox.information(self, "Nuovo Progetto", "Progetto inizializzato")
    
    def save_file(self):
        """Salva coordinate su file JSON"""
        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessuna coordinata da salvare")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Salva Coordinate Fili Fissi", "coordinate_fili_fissi.json", 
            "File JSON (*.json);;Tutti i file (*)"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(self.model.to_dict(), f, indent=2, ensure_ascii=False)
                QMessageBox.information(self, "Salvataggio Completato", 
                                      f"Coordinate salvate in:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Salvataggio", 
                                   f"Impossibile salvare il file:\n{e}")
    
    def open_file(self):
        """Apri file coordinate esistente"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "Apri Coordinate", "",
            "File JSON (*.json);;Tutti i file (*)"
        )

        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.model.from_dict(data)
                self.model.clear_undo_history()  # Pulisce cronologia quando carica file
                self.refresh_display()
                QMessageBox.information(self, "Caricamento Completato",
                                      f"Coordinate caricate da:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Caricamento",
                                   f"Impossibile aprire il file:\n{e}")
    
    def import_csv(self):
        """Importa coordinate da file CSV"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "Importa CSV", "",
            "File CSV (*.csv);;Tutti i file (*)"
        )

        if filename:
            try:
                imported = 0
                with open(filename, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        try:
                            x = float(row.get('X', row.get('x', 0)))
                            y = float(row.get('Y', row.get('y', 0)))
                            z = float(row.get('Z', row.get('z', 0)))
                            desc = row.get('Descrizione', row.get('Description', row.get('descrizione', '')))
                            level = row.get('Livello', row.get('Level', row.get('livello', '')))
                            self.model.add_point(x, y, z, desc, level)
                            imported += 1
                        except (ValueError, KeyError) as e:
                            continue

                self.refresh_display()
                QMessageBox.information(self, "Importazione Completata",
                                      f"{imported} punti importati da:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Importazione",
                                   f"Impossibile importare CSV:\n{e}")

    def export_csv(self):
        """Esporta coordinate in formato CSV"""
        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessuna coordinata da esportare")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Esporta CSV", "coordinate_fili_fissi.csv",
            "File CSV (*.csv);;Tutti i file (*)"
        )

        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write("ID,X,Y,Z,Livello,Descrizione\n")
                    for point in self.model.get_point_list():
                        f.write(f'{point.id},{point.x:.4f},{point.y:.4f},{point.z:.3f},"{point.level}","{point.description}"\n')
                QMessageBox.information(self, "Esportazione Completata",
                                      f"CSV esportato in:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Esportazione",
                                   f"Impossibile esportare CSV:\n{e}")

    def import_excel(self):
        """Importa coordinate da file Excel"""
        if not OPENPYXL_AVAILABLE:
            reply = QMessageBox.question(
                self, "Openpyxl Non Disponibile",
                "La libreria openpyxl non è installata.\n\nVuoi installare openpyxl ora?\n\nEsegui: pip install openpyxl",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                QMessageBox.information(self, "Installazione",
                                      "Apri un terminale ed esegui:\n\npip install openpyxl\n\nPoi riavvia l'applicazione.")
            return

        filename, _ = QFileDialog.getOpenFileName(
            self, "Importa Excel", "",
            "File Excel (*.xlsx);;Tutti i file (*)"
        )

        if filename:
            try:
                imported = self.model.import_excel(filename)
                self.refresh_display()
                QMessageBox.information(self, "Importazione Completata",
                                      f"{imported} punti importati da:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Importazione",
                                   f"Impossibile importare Excel:\n{e}")

    def export_excel(self):
        """Esporta coordinate in formato Excel"""
        if not OPENPYXL_AVAILABLE:
            reply = QMessageBox.question(
                self, "Openpyxl Non Disponibile",
                "La libreria openpyxl non è installata.\n\nVuoi installare openpyxl ora?\n\nEsegui: pip install openpyxl",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                QMessageBox.information(self, "Installazione",
                                      "Apri un terminale ed esegui:\n\npip install openpyxl\n\nPoi riavvia l'applicazione.")
            return

        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessuna coordinata da esportare")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Esporta Excel", "coordinate_fili_fissi.xlsx",
            "File Excel (*.xlsx);;Tutti i file (*)"
        )

        if filename:
            try:
                self.model.export_excel(filename)
                QMessageBox.information(self, "Esportazione Completata",
                                      f"Excel esportato in:\n{filename}\n\nIl file contiene 3 fogli:\n• Coordinate Fili Fissi\n• Info Progetto\n• Statistiche")
            except Exception as e:
                QMessageBox.critical(self, "Errore Esportazione",
                                   f"Impossibile esportare Excel:\n{e}")

    def import_dxf_ui(self):
        """Importa coordinate da file DXF"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "Importa DXF", "",
            "File DXF (*.dxf);;Tutti i file (*)"
        )

        if filename:
            try:
                imported = self.model.import_dxf(filename)
                self.refresh_display()

                if EZDXF_AVAILABLE:
                    mode_msg = "Importazione avanzata con ezdxf"
                else:
                    mode_msg = "Importazione con parser manuale"

                QMessageBox.information(self, "Importazione Completata",
                                      f"{imported} punti importati da:\n{filename}\n\n{mode_msg}\n\nEntità lette: POINT, TEXT, INSERT (blocchi)")
            except Exception as e:
                QMessageBox.critical(self, "Errore Importazione",
                                   f"Impossibile importare DXF:\n{e}\n\nSuggerimento: Installare ezdxf per import avanzato:\npip install ezdxf")

    def export_dxf(self):
        """Esporta coordinate in formato DXF (semplificato)"""
        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessuna coordinata da esportare")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Esporta DXF", "coordinate_fili_fissi.dxf",
            "File DXF (*.dxf);;Tutti i file (*)"
        )

        if filename:
            try:
                # Export DXF semplificato (senza libreria ezdxf)
                with open(filename, 'w', encoding='utf-8') as f:
                    # Header DXF minimale
                    f.write("  0\nSECTION\n  2\nENTITIES\n")

                    # Aggiungi punti come POINT entities
                    for point in self.model.get_point_list():
                        f.write("  0\nPOINT\n")
                        f.write(f" 10\n{point.x:.4f}\n")
                        f.write(f" 20\n{point.y:.4f}\n")
                        f.write(f" 30\n{point.z:.3f}\n")
                        # Aggiungi testo come descrizione
                        if point.description:
                            f.write("  0\nTEXT\n")
                            f.write(f" 10\n{point.x:.4f}\n")
                            f.write(f" 20\n{point.y:.4f}\n")
                            f.write(f" 30\n{point.z:.3f}\n")
                            f.write(f"  1\n{point.description}\n")
                            f.write(" 40\n0.5\n")  # Text height

                    f.write("  0\nENDSEC\n  0\nEOF\n")

                QMessageBox.information(self, "Esportazione Completata",
                                      f"DXF esportato in:\n{filename}\n\nNota: formato DXF semplificato")
            except Exception as e:
                QMessageBox.critical(self, "Errore Esportazione",
                                   f"Impossibile esportare DXF:\n{e}")

    def generate_pdf_report_ui(self):
        """Genera report PDF completo"""
        if not REPORTLAB_AVAILABLE:
            reply = QMessageBox.question(
                self, "Reportlab Non Disponibile",
                "La libreria reportlab non è installata.\n\nVuoi installare reportlab ora?\n\nEsegui: pip install reportlab",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                QMessageBox.information(self, "Installazione",
                                      "Apri un terminale ed esegui:\n\npip install reportlab\n\nPoi riavvia l'applicazione.")
            return

        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessuna coordinata da esportare")
            return

        # Verifica che ci siano metadati minimi
        if not self.model.metadata.get('project_name'):
            reply = QMessageBox.question(
                self, "Informazioni Progetto Mancanti",
                "Nessun nome progetto specificato.\n\nVuoi inserire le informazioni del progetto prima di generare il report?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.edit_metadata()

        filename, _ = QFileDialog.getSaveFileName(
            self, "Genera Report PDF", "report_coordinate_fili_fissi.pdf",
            "File PDF (*.pdf);;Tutti i file (*)"
        )

        if filename:
            try:
                self.model.generate_pdf_report(filename)
                QMessageBox.information(self, "Report Generato",
                                      f"Report PDF generato con successo:\n{filename}\n\nIl report contiene:\n• Frontespizio con info progetto\n• Statistiche coordinate\n• Livelli definiti\n• Tabella completa coordinate")
            except Exception as e:
                QMessageBox.critical(self, "Errore Generazione Report",
                                   f"Impossibile generare report PDF:\n{e}")

    def edit_metadata(self):
        """Modifica metadata progetto"""
        dialog = MetadataDialog(self.model.metadata, self)
        if dialog.exec_() == QDialog.Accepted:
            updated_metadata = dialog.get_metadata()
            self.model.metadata.update(updated_metadata)
            self.model._mark_modified()
            self.refresh_display()
            QMessageBox.information(self, "Metadata Aggiornato",
                                  "Le informazioni del progetto sono state aggiornate")

    def duplicate_selected_point(self):
        """Duplica punto selezionato"""
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Attenzione", "Seleziona un punto da duplicare")
            return

        id_item = self.table.item(current_row, 0)
        if id_item:
            point_id = int(id_item.text())

            # Chiedi offset (opzionale)
            offset_x, ok1 = QInputDialog.getDouble(self, "Offset X", "Offset X (m):", 0.0, -10000, 10000, 4)
            if not ok1:
                return
            offset_y, ok2 = QInputDialog.getDouble(self, "Offset Y", "Offset Y (m):", 0.0, -10000, 10000, 4)
            if not ok2:
                return
            offset_z, ok3 = QInputDialog.getDouble(self, "Offset Z", "Offset Z (m):", 0.0, -1000, 1000, 3)
            if not ok3:
                return

            new_id = self.model.duplicate_point(point_id, offset_x, offset_y, offset_z)
            if new_id:
                self.refresh_display()
                QMessageBox.information(self, "Punto Duplicato",
                                      f"Punto {point_id} duplicato come punto {new_id}")

    def translate_coordinates(self):
        """Trasla tutte le coordinate"""
        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessun punto da traslare")
            return

        dx, ok1 = QInputDialog.getDouble(self, "Traslazione", "Offset X (m):", 0.0, -10000, 10000, 4)
        if not ok1:
            return
        dy, ok2 = QInputDialog.getDouble(self, "Traslazione", "Offset Y (m):", 0.0, -10000, 10000, 4)
        if not ok2:
            return
        dz, ok3 = QInputDialog.getDouble(self, "Traslazione", "Offset Z (m):", 0.0, -1000, 1000, 3)
        if not ok3:
            return

        if dx == 0 and dy == 0 and dz == 0:
            QMessageBox.information(self, "Traslazione", "Nessun offset applicato")
            return

        reply = QMessageBox.question(
            self, 'Conferma Traslazione',
            f'Traslare tutti i {len(self.model.points)} punti di:\nΔX={dx:.4f}m, ΔY={dy:.4f}m, ΔZ={dz:.3f}m?',
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.model.translate_all(dx, dy, dz)
            self.refresh_display()
            QMessageBox.information(self, "Traslazione Completata",
                                  f"Tutti i punti sono stati traslati")

    def find_duplicate_points(self):
        """Trova punti duplicati"""
        if not self.model.points:
            QMessageBox.information(self, "Trova Duplicati", "Nessun punto da analizzare")
            return

        tolerance, ok = QInputDialog.getDouble(
            self, "Tolleranza", "Tolleranza distanza (m):", 0.001, 0.0001, 1.0, 4
        )
        if not ok:
            return

        duplicates = self.model.find_duplicates(tolerance)

        if not duplicates:
            QMessageBox.information(self, "Trova Duplicati",
                                  f"Nessun punto duplicato trovato (tolleranza: {tolerance:.4f}m)")
        else:
            msg = f"Trovati {len(duplicates)} coppie di punti duplicati:\n\n"
            for id1, id2, dist in duplicates[:10]:  # Mostra max 10
                p1 = self.model.points[id1]
                p2 = self.model.points[id2]
                msg += f"• Punto {id1} e {id2} - distanza: {dist:.4f}m\n"
            if len(duplicates) > 10:
                msg += f"\n... e altri {len(duplicates)-10} duplicati"

            QMessageBox.warning(self, "Punti Duplicati Trovati", msg)

    def calculate_distance_dialog(self):
        """Calcola distanza tra due punti"""
        if len(self.model.points) < 2:
            QMessageBox.warning(self, "Attenzione", "Servono almeno 2 punti")
            return

        # Lista ID punti
        point_ids = [str(p.id) for p in self.model.get_point_list()]

        # Selezione punto 1
        id1_str, ok1 = QInputDialog.getItem(
            self, "Punto 1", "Seleziona primo punto (ID):", point_ids, 0, False
        )
        if not ok1:
            return
        id1 = int(id1_str)

        # Selezione punto 2
        id2_str, ok2 = QInputDialog.getItem(
            self, "Punto 2", "Seleziona secondo punto (ID):", point_ids, 0, False
        )
        if not ok2:
            return
        id2 = int(id2_str)

        if id1 == id2:
            QMessageBox.warning(self, "Attenzione", "Seleziona due punti diversi")
            return

        result = self.model.calculate_distance(id1, id2)
        if result:
            p1 = self.model.points[id1]
            p2 = self.model.points[id2]

            msg = f"""DISTANZA TRA PUNTI

Punto {id1}: ({p1.x:.4f}, {p1.y:.4f}, {p1.z:.3f}) - {p1.description}
Punto {id2}: ({p2.x:.4f}, {p2.y:.4f}, {p2.z:.3f}) - {p2.description}

DIFFERENZE
ΔX: {result['dx']:.4f} m
ΔY: {result['dy']:.4f} m
ΔZ: {result['dz']:.3f} m

DISTANZE
Distanza 2D (planimetrica): {result['distance_2d']:.4f} m
Distanza 3D (spaziale): {result['distance_3d']:.4f} m
"""
            QMessageBox.information(self, "Calcolo Distanza", msg)

    def assign_level_dialog(self):
        """Assegna livello a punti selezionati"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Attenzione", "Seleziona almeno un punto")
            return

        level, ok = QInputDialog.getText(self, "Assegna Livello",
                                         "Livello (es: PT, P1, P2):")
        if ok and level:
            point_ids = []
            for row in selected_rows:
                id_item = self.table.item(row.row(), 0)
                if id_item:
                    point_ids.append(int(id_item.text()))

            self.model.assign_level_to_points(point_ids, level)
            self.refresh_display()
            QMessageBox.information(self, "Livello Assegnato",
                                  f"Livello '{level}' assegnato a {len(point_ids)} punti")

    def copy_level_dialog(self):
        """Copia piano su altro livello"""
        levels = self.model.get_levels()
        if not levels:
            QMessageBox.warning(self, "Attenzione", "Nessun livello definito")
            return

        source, ok1 = QInputDialog.getItem(self, "Copia Livello",
                                           "Livello sorgente:", levels, 0, False)
        if not ok1:
            return

        target, ok2 = QInputDialog.getText(self, "Copia Livello",
                                           "Nuovo livello destinazione:")
        if not ok2 or not target:
            return

        offset_z, ok3 = QInputDialog.getDouble(self, "Offset Z",
                                                "Offset verticale (m):", 3.5, -100, 100, 2)
        if not ok3:
            return

        copied = self.model.copy_level(source, target, offset_z)
        self.refresh_display()
        QMessageBox.information(self, "Livello Copiato",
                              f"{copied} punti copiati da '{source}' a '{target}'")

    def manage_levels_dialog(self):
        """Gestione livelli"""
        levels = self.model.get_levels()
        if not levels:
            QMessageBox.information(self, "Gestione Livelli", "Nessun livello definito")
            return

        msg = "LIVELLI DEFINITI:\n\n"
        for level in levels:
            points_count = len(self.model.get_points_by_level(level))
            msg += f"• {level}: {points_count} punti\n"

        QMessageBox.information(self, "Gestione Livelli", msg)

    def generate_grid_dialog(self):
        """Genera griglia regolare"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Genera Griglia Regolare")
        layout = QGridLayout()

        # Input parametri
        layout.addWidget(QLabel("X inizio (m):"), 0, 0)
        x_start = QDoubleSpinBox()
        x_start.setRange(-1000, 1000)
        x_start.setValue(0.0)
        x_start.setDecimals(2)
        layout.addWidget(x_start, 0, 1)

        layout.addWidget(QLabel("X fine (m):"), 0, 2)
        x_end = QDoubleSpinBox()
        x_end.setRange(-1000, 1000)
        x_end.setValue(20.0)
        x_end.setDecimals(2)
        layout.addWidget(x_end, 0, 3)

        layout.addWidget(QLabel("Numero assi X:"), 1, 0)
        x_count = QSpinBox()
        x_count.setRange(1, 26)
        x_count.setValue(5)
        layout.addWidget(x_count, 1, 1)

        layout.addWidget(QLabel("Y inizio (m):"), 2, 0)
        y_start = QDoubleSpinBox()
        y_start.setRange(-1000, 1000)
        y_start.setValue(0.0)
        y_start.setDecimals(2)
        layout.addWidget(y_start, 2, 1)

        layout.addWidget(QLabel("Y fine (m):"), 2, 2)
        y_end = QDoubleSpinBox()
        y_end.setRange(-1000, 1000)
        y_end.setValue(15.0)
        y_end.setDecimals(2)
        layout.addWidget(y_end, 2, 3)

        layout.addWidget(QLabel("Numero assi Y:"), 3, 0)
        y_count = QSpinBox()
        y_count.setRange(1, 50)
        y_count.setValue(4)
        layout.addWidget(y_count, 3, 1)

        layout.addWidget(QLabel("Quota Z (m):"), 4, 0)
        z_value = QDoubleSpinBox()
        z_value.setRange(-100, 100)
        z_value.setValue(0.0)
        z_value.setDecimals(3)
        layout.addWidget(z_value, 4, 1)

        layout.addWidget(QLabel("Livello:"), 4, 2)
        level_input = QLineEdit()
        level_input.setText("PT")
        layout.addWidget(level_input, 4, 3)

        layout.addWidget(QLabel("Prefisso (opz):"), 5, 0)
        prefix = QLineEdit()
        prefix.setPlaceholderText("es: F1-")
        layout.addWidget(prefix, 5, 1)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons, 6, 0, 1, 4)

        dialog.setLayout(layout)

        if dialog.exec_() == QDialog.Accepted:
            generated = self.model.generate_grid(
                x_start.value(), x_end.value(), x_count.value(),
                y_start.value(), y_end.value(), y_count.value(),
                z_value.value(), level_input.text(), prefix.text()
            )
            self.refresh_display()
            QMessageBox.information(self, "Griglia Generata",
                                  f"{generated} punti generati")

    def generate_floors_dialog(self):
        """Genera interpiano automatico"""
        levels = self.model.get_levels()
        if not levels:
            QMessageBox.warning(self, "Attenzione", "Nessun livello base definito")
            return

        source, ok1 = QInputDialog.getItem(self, "Interpiano Automatico",
                                           "Livello base da duplicare:", levels, 0, False)
        if not ok1:
            return

        floors, ok2 = QInputDialog.getInt(self, "Numero Piani",
                                          "Quanti piani superiori generare:", 3, 1, 20)
        if not ok2:
            return

        height, ok3 = QInputDialog.getDouble(self, "Altezza Interpiano",
                                             "Altezza interpiano (m):", 3.5, 0.1, 10.0, 2)
        if not ok3:
            return

        total_copied = 0
        for i in range(1, floors + 1):
            target_level = f"P{i}"
            offset_z = i * height
            copied = self.model.copy_level(source, target_level, offset_z)
            total_copied += copied

        self.refresh_display()
        QMessageBox.information(self, "Interpiano Generato",
                              f"{total_copied} punti generati su {floors} piani")

    def select_by_area_dialog(self):
        """Seleziona punti per area"""
        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessun punto disponibile")
            return

        bounds = self.model.get_bounds()
        dialog = QDialog(self)
        dialog.setWindowTitle("Seleziona per Area")
        layout = QGridLayout()

        layout.addWidget(QLabel("X min (m):"), 0, 0)
        x_min = QDoubleSpinBox()
        x_min.setRange(-10000, 10000)
        x_min.setValue(bounds['x_min'])
        x_min.setDecimals(4)
        layout.addWidget(x_min, 0, 1)

        layout.addWidget(QLabel("X max (m):"), 0, 2)
        x_max = QDoubleSpinBox()
        x_max.setRange(-10000, 10000)
        x_max.setValue(bounds['x_max'])
        x_max.setDecimals(4)
        layout.addWidget(x_max, 0, 3)

        layout.addWidget(QLabel("Y min (m):"), 1, 0)
        y_min = QDoubleSpinBox()
        y_min.setRange(-10000, 10000)
        y_min.setValue(bounds['y_min'])
        y_min.setDecimals(4)
        layout.addWidget(y_min, 1, 1)

        layout.addWidget(QLabel("Y max (m):"), 1, 2)
        y_max = QDoubleSpinBox()
        y_max.setRange(-10000, 10000)
        y_max.setValue(bounds['y_max'])
        y_max.setDecimals(4)
        layout.addWidget(y_max, 1, 3)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons, 2, 0, 1, 4)

        dialog.setLayout(layout)

        if dialog.exec_() == QDialog.Accepted:
            selected = self.model.select_points_by_area(
                x_min.value(), x_max.value(), y_min.value(), y_max.value()
            )
            QMessageBox.information(self, "Selezione per Area",
                                  f"{len(selected)} punti selezionati:\n{selected[:20]}")

    def select_by_z_dialog(self):
        """Seleziona punti per quota Z"""
        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessun punto disponibile")
            return

        z_value, ok1 = QInputDialog.getDouble(self, "Seleziona per Quota",
                                               "Quota Z (m):", 0.0, -100, 100, 3)
        if not ok1:
            return

        tolerance, ok2 = QInputDialog.getDouble(self, "Tolleranza",
                                                 "Tolleranza (m):", 0.001, 0.0001, 1.0, 4)
        if not ok2:
            return

        selected = self.model.select_points_by_z(z_value, tolerance)
        QMessageBox.information(self, "Selezione per Quota",
                              f"{len(selected)} punti a quota Z={z_value:.3f}m:\n{selected[:20]}")

    def translate_selection_dialog(self):
        """Trasla punti selezionati"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Attenzione", "Seleziona almeno un punto")
            return

        point_ids = []
        for row in selected_rows:
            id_item = self.table.item(row.row(), 0)
            if id_item:
                point_ids.append(int(id_item.text()))

        dx, ok1 = QInputDialog.getDouble(self, "Traslazione", "Offset X (m):", 0.0, -10000, 10000, 4)
        if not ok1:
            return
        dy, ok2 = QInputDialog.getDouble(self, "Traslazione", "Offset Y (m):", 0.0, -10000, 10000, 4)
        if not ok2:
            return
        dz, ok3 = QInputDialog.getDouble(self, "Traslazione", "Offset Z (m):", 0.0, -1000, 1000, 3)
        if not ok3:
            return

        self.model.translate_points(point_ids, dx, dy, dz)
        self.refresh_display()
        QMessageBox.information(self, "Traslazione Completata",
                              f"{len(point_ids)} punti traslati")

    def delete_selection(self):
        """Elimina punti selezionati"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Attenzione", "Seleziona almeno un punto")
            return

        point_ids = []
        for row in selected_rows:
            id_item = self.table.item(row.row(), 0)
            if id_item:
                point_ids.append(int(id_item.text()))

        reply = QMessageBox.question(
            self, 'Conferma Eliminazione',
            f'Eliminare {len(point_ids)} punti selezionati?',
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.model.delete_points(point_ids)
            self.refresh_display()
            QMessageBox.information(self, "Eliminazione Completata",
                                  f"{len(point_ids)} punti eliminati")

    def mirror_points_dialog(self):
        """Specchia punti selezionati"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Attenzione", "Seleziona almeno un punto")
            return

        point_ids = []
        for row in selected_rows:
            id_item = self.table.item(row.row(), 0)
            if id_item:
                point_ids.append(int(id_item.text()))

        axis, ok1 = QInputDialog.getItem(self, "Specchiamento",
                                         "Asse di specchiamento:", ['X', 'Y', 'Z'], 0, False)
        if not ok1:
            return

        value, ok2 = QInputDialog.getDouble(self, "Valore Asse",
                                            f"Valore asse {axis} (m):", 0.0, -1000, 1000, 3)
        if not ok2:
            return

        self.model.mirror_points(point_ids, axis.lower(), value)
        self.refresh_display()
        QMessageBox.information(self, "Specchiamento Completato",
                              f"{len(point_ids)} punti specchiati rispetto a {axis}={value:.3f}m")

    def snap_to_grid_dialog(self):
        """Snap punti a griglia"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Attenzione", "Seleziona almeno un punto")
            return

        point_ids = []
        for row in selected_rows:
            id_item = self.table.item(row.row(), 0)
            if id_item:
                point_ids.append(int(id_item.text()))

        grid_size, ok = QInputDialog.getDouble(self, "Snap to Grid",
                                                "Dimensione griglia (m):", 0.5, 0.01, 10.0, 3)
        if not ok:
            return

        self.model.snap_to_grid(point_ids, grid_size)
        self.refresh_display()
        QMessageBox.information(self, "Snap Completato",
                              f"{len(point_ids)} punti arrotondati a griglia {grid_size}m")

    def show_plot_window(self):
        """Mostra finestra visualizzazione grafica"""
        if not self.model.points:
            QMessageBox.warning(self, "Attenzione", "Nessun punto da visualizzare")
            return

        if not MATPLOTLIB_AVAILABLE:
            reply = QMessageBox.question(
                self, "Matplotlib Non Disponibile",
                "Matplotlib non è installato.\n\nVuoi installarlo ora?\n\nEsegui: pip install matplotlib",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                QMessageBox.information(self, "Installazione",
                                      "Apri un terminale ed esegui:\n\npip install matplotlib\n\nPoi riavvia l'applicazione.")
            return

        dialog = PlotDialog(self.model, self)
        dialog.exec_()

    def show_statistics(self):
        """Mostra statistiche complete del progetto"""
        points = list(self.model.points.values())
        if not points:
            QMessageBox.information(self, "Statistiche", "Nessuna coordinata inserita")
            return
        
        bounds = self.model.get_bounds()
        
        # Calcola distanze
        x_range = bounds['x_max'] - bounds['x_min']
        y_range = bounds['y_max'] - bounds['y_min']
        z_range = bounds['z_max'] - bounds['z_min']
        
        stats_text = f"""STATISTICHE COORDINATE FILI FISSI

Progetto: {self.model.metadata.get('project_name', 'Non specificato')}
Architetto: {self.model.metadata.get('author', 'N/A')}

PUNTI COORDINATI
Totale punti: {len(points)}

ESTENSIONI PLANIMETRICHE
X min: {bounds['x_min']:.4f} m    X max: {bounds['x_max']:.4f} m    Δ: {x_range:.4f} m
Y min: {bounds['y_min']:.4f} m    Y max: {bounds['y_max']:.4f} m    Δ: {y_range:.4f} m
Z min: {bounds['z_min']:.3f} m     Z max: {bounds['z_max']:.3f} m     Δ: {z_range:.3f} m

INFORMAZIONI FILE
Creato: {self.model.metadata.get('created', 'N/A')[:19]}
Modificato: {self.model.metadata.get('modified', 'N/A')[:19]}
"""
        
        QMessageBox.information(self, "Statistiche Progetto", stats_text)
    
    def show_coordinate_info(self):
        """Mostra informazioni dettagliate sulle coordinate"""
        points = list(self.model.points.values())
        if not points:
            QMessageBox.information(self, "Info Coordinate", "Nessuna coordinata disponibile")
            return
        
        # Trova punti estremi
        x_coords = [p.x for p in points]
        y_coords = [p.y for p in points]
        
        x_min_point = min(points, key=lambda p: p.x)
        x_max_point = max(points, key=lambda p: p.x)
        y_min_point = min(points, key=lambda p: p.y)
        y_max_point = max(points, key=lambda p: p.y)
        
        info_text = f"""INFORMAZIONI COORDINATE

PUNTI ESTREMI
X minima: Punto {x_min_point.id} ({x_min_point.x:.4f}, {x_min_point.y:.4f}) - {x_min_point.description}
X massima: Punto {x_max_point.id} ({x_max_point.x:.4f}, {x_max_point.y:.4f}) - {x_max_point.description}

Y minima: Punto {y_min_point.id} ({y_min_point.x:.4f}, {y_min_point.y:.4f}) - {y_min_point.description}
Y massima: Punto {y_max_point.id} ({y_max_point.x:.4f}, {y_max_point.y:.4f}) - {y_max_point.description}

COORDINATE CENTRO BARICENTRICO
X centro: {sum(x_coords)/len(x_coords):.4f} m
Y centro: {sum(y_coords)/len(y_coords):.4f} m
"""
        
        QMessageBox.information(self, "Informazioni Coordinate", info_text)
    
    def clear_all_points(self):
        """Elimina tutti i punti"""
        if not self.model.points:
            QMessageBox.information(self, "Elimina Tutto", "Nessun punto da eliminare")
            return
        
        reply = QMessageBox.question(
            self, 'Conferma Eliminazione', 
            f'Eliminare tutti i {len(self.model.points)} punti?',
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.model.clear()
            self.refresh_display()
            QMessageBox.information(self, "Eliminazione Completata", "Tutti i punti sono stati eliminati")
    
    def show_about(self):
        """Informazioni sull'applicazione"""
        matplotlib_status = "✓ Disponibile" if MATPLOTLIB_AVAILABLE else "✗ Non installato"
        openpyxl_status = "✓ Disponibile" if OPENPYXL_AVAILABLE else "✗ Non installato"
        reportlab_status = "✓ Disponibile" if REPORTLAB_AVAILABLE else "✗ Non installato"
        ezdxf_status = "✓ Disponibile" if EZDXF_AVAILABLE else "✗ Non installato (fallback a parser manuale)"

        about_text = f"""INPUT COORDINATE FILI FISSI

Applicazione professionale avanzata per l'inserimento e gestione
delle coordinate spaziali per la progettazione di fili fissi strutturali.

Sviluppato per:
Arch. Michelangelo Bartolotta
Ordine Architetti Agrigento n. 1557

Funzionalità Base:
• Input rapido coordinate X, Y, Z con livelli
• Modifica diretta in tabella (6 colonne)
• Import/Export: JSON, CSV, Excel (.xlsx), DXF
• Calcolo distanze tra punti (2D/3D)
• Rilevamento punti duplicati
• Undo/Redo completo (Ctrl+Z / Ctrl+Y) con 50 livelli

Gestione Livelli/Piani:
• Assegnazione livelli a punti
• Copia piano con offset verticale
• Interpiano automatico multipiano
• Filtro visualizzazione per livello

Generatori Automatici:
• Griglia regolare con nomenclatura (A1, A2, B1...)
• Generazione multipiano

Selezione Avanzata:
• Selezione per area rettangolare
• Selezione per quota Z
• Operazioni su selezione multipla

Trasformazioni:
• Traslazione punti selezionati
• Specchiamento rispetto ad assi
• Snap to grid parametrico
• Duplicazione con offset

Visualizzazione:
• Vista grafica 2D/3D interattiva
• Plot planimetria (XY), sezioni (XZ, YZ)
• Vista 3D navigabile
• Colori per livelli, etichette
• Export immagini PNG ad alta risoluzione

Export Excel:
• File multi-foglio con formattazione professionale
• Foglio Coordinate con stili e bordi
• Foglio Info Progetto con metadata
• Foglio Statistiche automatiche

Report PDF (Ctrl+P):
• Frontespizio con informazioni progetto
• Statistiche estensioni coordinate
• Tabella livelli definiti
• Elenco completo coordinate con formattazione
• Layout professionale A4 pronto per stampa

Import/Export DXF:
• Import DXF con ezdxf (opzionale) o parser manuale
• Lettura entità POINT, TEXT, INSERT (blocchi)
• Layers come livelli
• Export DXF semplificato

Annotazioni Grafiche:
• Linee tra punti (stili: continua, tratteggiata, punteggiata)
• Quote/dimensioni con offset configurabile
• Frecce e testo automatico con distanza
• Rendering in tutte le viste 2D (XY, XZ, YZ)

Librerie Opzionali:
Matplotlib: {matplotlib_status}
Openpyxl: {openpyxl_status}
Reportlab: {"✓ Disponibile" if REPORTLAB_AVAILABLE else "✗ Non installato"}
Ezdxf: {"✓ Disponibile" if EZDXF_AVAILABLE else "✗ Non installato (fallback a parser manuale)"}

Versione: 5.3 - DXF Import & Annotations
"""
        QMessageBox.about(self, "Informazioni", about_text)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName("Input Coordinate Fili Fissi")
    app.setApplicationVersion("5.3")
    
    # Stile applicazione
    app.setStyleSheet("""
        QGroupBox {
            font-weight: bold;
            border: 2px solid #cccccc;
            border-radius: 5px;
            margin-top: 1ex;
            padding-top: 10px;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px 0 5px;
        }
        QPushButton {
            padding: 5px 10px;
            border-radius: 3px;
        }
        QPushButton:default {
            background-color: #007acc;
            color: white;
            border: none;
        }
        QDoubleSpinBox {
            padding: 3px;
        }
    """)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())