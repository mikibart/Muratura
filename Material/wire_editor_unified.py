#!/usr/bin/env python3
"""
Wire Editor Unified v6.1
Applicazione integrata per gestione coordinate fili fissi

Combina funzionalità di:
- wire_editor.py (input tabellare, import/export, reports)
- wire_editor1.py (canvas grafico, topologie, disegno interattivo)

Arch. Michelangelo Bartolotta
Ordine Architetti Agrigento n. 1557
"""

import sys
import json
import math
import csv
from datetime import datetime
from enum import Enum
from copy import deepcopy
from typing import Dict, List, Optional, Tuple, Any

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QDoubleSpinBox, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QGroupBox, QGridLayout, QFileDialog, QMessageBox,
    QHeaderView, QAbstractItemView, QDialog, QTextEdit, QDialogButtonBox,
    QInputDialog, QComboBox, QSpinBox, QCheckBox, QTabWidget, QToolBar,
    QAction, QSplitter, QScrollArea, QDockWidget
)
from PyQt5.QtCore import Qt, QPointF, QRectF, pyqtSignal, QTimer
from PyQt5.QtGui import (
    QPainter, QPen, QBrush, QColor, QFont, QIcon, QCursor,
    QPixmap, QPalette
)

# Matplotlib imports (opzionali)
try:
    import matplotlib
    matplotlib.use('Qt5Agg')
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# Openpyxl imports (opzionali)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Reportlab imports (opzionali)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Ezdxf imports (opzionali)
try:
    import ezdxf
    EZDXF_AVAILABLE = True
except ImportError:
    EZDXF_AVAILABLE = False


# ============================================================================
# ENUMERAZIONI E COSTANTI
# ============================================================================

class Direction(Enum):
    """Direzione per edges"""
    NORTH = "north"
    SOUTH = "south"
    EAST = "east"
    WEST = "west"
    NONE = "none"


class TopologyRole(Enum):
    """Ruolo topologico del nodo"""
    CENTER = "center"
    CORNER = "corner"
    EDGE_MIDPOINT = "edge"
    FREE = "free"


class EdgeType(Enum):
    """Tipo di collegamento"""
    HORIZONTAL = "horizontal"
    VERTICAL = "vertical"
    DIAGONAL = "diagonal"
    FREE = "free"


# ============================================================================
# DATA MODEL UNIFICATO
# ============================================================================

class Node:
    """Nodo unificato - combina Point (wire_editor.py) e Node (wire_editor1.py)"""

    def __init__(self, node_id: int, x: float, y: float, z: float = 0.0,
                 description: str = "", level: str = "",
                 topology_role: TopologyRole = TopologyRole.FREE):
        self.id = node_id
        self.x = x
        self.y = y
        self.z = z
        self.description = description
        self.level = level  # Livello/Piano (es: "PT", "P1", "P2")

        # Proprietà topologiche (da wire_editor1.py)
        self.topology_role = topology_role
        self.snap_radius = 0.15  # metri
        self.color = QColor(100, 150, 200)
        self.selected = False

    def to_dict(self) -> Dict[str, Any]:
        """Serializzazione per JSON"""
        return {
            'id': self.id,
            'x': self.x,
            'y': self.y,
            'z': self.z,
            'description': self.description,
            'level': self.level,
            'topology_role': self.topology_role.value if self.topology_role else TopologyRole.FREE.value,
            'snap_radius': self.snap_radius,
            'color': self.color.name()
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'Node':
        """Deserializzazione da JSON"""
        node = cls(
            data['id'],
            data['x'],
            data['y'],
            data.get('z', 0.0),
            data.get('description', ''),
            data.get('level', ''),
            TopologyRole(data.get('topology_role', 'free'))
        )
        node.snap_radius = data.get('snap_radius', 0.15)
        if 'color' in data:
            node.color = QColor(data['color'])
        return node

    def distance_to(self, other: 'Node') -> float:
        """Distanza 2D ad altro nodo"""
        return math.sqrt((self.x - other.x)**2 + (self.y - other.y)**2)

    def distance_3d_to(self, other: 'Node') -> float:
        """Distanza 3D ad altro nodo"""
        return math.sqrt((self.x - other.x)**2 + (self.y - other.y)**2 + (self.z - other.z)**2)


class Edge:
    """Collegamento tra nodi (da wire_editor1.py)"""

    def __init__(self, edge_id: int, node1_id: int, node2_id: int,
                 edge_type: EdgeType = EdgeType.FREE):
        self.id = edge_id
        self.node1_id = node1_id
        self.node2_id = node2_id
        self.edge_type = edge_type
        self.color = QColor(80, 80, 80)
        self.width = 2
        self.selected = False

    def to_dict(self) -> Dict[str, Any]:
        """Serializzazione per JSON"""
        return {
            'id': self.id,
            'node1_id': self.node1_id,
            'node2_id': self.node2_id,
            'edge_type': self.edge_type.value,
            'color': self.color.name(),
            'width': self.width
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'Edge':
        """Deserializzazione da JSON"""
        edge = cls(
            data['id'],
            data['node1_id'],
            data['node2_id'],
            EdgeType(data.get('edge_type', 'free'))
        )
        if 'color' in data:
            edge.color = QColor(data['color'])
        edge.width = data.get('width', 2)
        return edge


class Annotation:
    """Annotazione grafica (da wire_editor.py)"""

    def __init__(self, annotation_type: str, color: str = 'red', linewidth: int = 2):
        self.type = annotation_type
        self.color = color
        self.linewidth = linewidth

    def to_dict(self) -> Dict[str, Any]:
        return {
            'type': self.type,
            'color': self.color,
            'linewidth': self.linewidth
        }


class LineAnnotation(Annotation):
    """Linea tra due nodi"""

    def __init__(self, node1_id: int, node2_id: int, color: str = 'red',
                 linewidth: int = 2, style: str = 'solid'):
        super().__init__('line', color, linewidth)
        self.node1_id = node1_id
        self.node2_id = node2_id
        self.style = style  # 'solid', 'dashed', 'dotted'

    def to_dict(self) -> Dict[str, Any]:
        data = super().to_dict()
        data.update({
            'node1_id': self.node1_id,
            'node2_id': self.node2_id,
            'style': self.style
        })
        return data

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'LineAnnotation':
        return cls(
            data['node1_id'],
            data['node2_id'],
            data.get('color', 'red'),
            data.get('linewidth', 2),
            data.get('style', 'solid')
        )


class DimensionAnnotation(Annotation):
    """Quota/dimensione tra due nodi"""

    def __init__(self, node1_id: int, node2_id: int, offset: float = 0.5,
                 color: str = 'blue', show_text: bool = True):
        super().__init__('dimension', color, 1)
        self.node1_id = node1_id
        self.node2_id = node2_id
        self.offset = offset
        self.show_text = show_text

    def to_dict(self) -> Dict[str, Any]:
        data = super().to_dict()
        data.update({
            'node1_id': self.node1_id,
            'node2_id': self.node2_id,
            'offset': self.offset,
            'show_text': self.show_text
        })
        return data

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'DimensionAnnotation':
        return cls(
            data['node1_id'],
            data['node2_id'],
            data.get('offset', 0.5),
            data.get('color', 'blue'),
            data.get('show_text', True)
        )


class UnifiedModel:
    """Model unificato - gestisce nodi, edges, annotazioni"""

    def __init__(self):
        self.nodes: Dict[int, Node] = {}
        self.edges: Dict[int, Edge] = {}
        self.annotations: List[Annotation] = []

        self.next_node_id = 1
        self.next_edge_id = 1

        self.metadata = {
            "author": "Arch. Michelangelo Bartolotta",
            "registration": "Ordine Architetti Agrigento n. 1557",
            "project_type": "Coordinate Fili Fissi",
            "created": datetime.now().isoformat(),
            "modified": datetime.now().isoformat(),
            "project_name": "",
            "location": "",
            "notes": ""
        }

        # Undo/Redo
        self.undo_stack: List['Command'] = []
        self.redo_stack: List['Command'] = []
        self.max_undo_levels = 50

    # === Gestione Nodi ===

    def add_node(self, x: float, y: float, z: float = 0.0,
                 description: str = "", level: str = "",
                 topology_role: TopologyRole = TopologyRole.FREE) -> int:
        """Aggiunge un nodo (interno - usare execute_command per undo/redo)"""
        if not description:
            description = f"Node {self.next_node_id}"

        node = Node(self.next_node_id, x, y, z, description, level, topology_role)
        self.nodes[self.next_node_id] = node
        node_id = self.next_node_id
        self.next_node_id += 1
        self._mark_modified()
        return node_id

    def update_node(self, node_id: int, **kwargs) -> bool:
        """Aggiorna un nodo"""
        if node_id in self.nodes:
            node = self.nodes[node_id]
            for key, value in kwargs.items():
                if hasattr(node, key):
                    setattr(node, key, value)
            self._mark_modified()
            return True
        return False

    def delete_node(self, node_id: int) -> bool:
        """Elimina un nodo e tutti gli edge connessi"""
        if node_id in self.nodes:
            # Elimina edges connessi
            edges_to_delete = [eid for eid, edge in self.edges.items()
                             if edge.node1_id == node_id or edge.node2_id == node_id]
            for eid in edges_to_delete:
                del self.edges[eid]

            # Elimina nodo
            del self.nodes[node_id]
            self._mark_modified()
            return True
        return False

    def get_node_list(self) -> List[Node]:
        """Restituisce lista ordinata nodi"""
        return sorted(self.nodes.values(), key=lambda n: n.id)

    # === Gestione Edges ===

    def add_edge(self, node1_id: int, node2_id: int,
                 edge_type: EdgeType = EdgeType.FREE) -> Optional[int]:
        """Aggiunge un edge tra due nodi"""
        if node1_id not in self.nodes or node2_id not in self.nodes:
            return None

        # Verifica se edge già esiste
        for edge in self.edges.values():
            if (edge.node1_id == node1_id and edge.node2_id == node2_id) or \
               (edge.node1_id == node2_id and edge.node2_id == node1_id):
                return None  # Edge già esiste

        edge = Edge(self.next_edge_id, node1_id, node2_id, edge_type)
        self.edges[self.next_edge_id] = edge
        edge_id = self.next_edge_id
        self.next_edge_id += 1
        self._mark_modified()
        return edge_id

    def delete_edge(self, edge_id: int) -> bool:
        """Elimina un edge"""
        if edge_id in self.edges:
            del self.edges[edge_id]
            self._mark_modified()
            return True
        return False

    # === Utility ===

    def get_levels(self) -> List[str]:
        """Restituisce lista livelli unici ordinati"""
        levels = set(n.level for n in self.nodes.values() if n.level)
        return sorted(levels)

    def get_nodes_by_level(self, level: str) -> List[Node]:
        """Restituisce nodi di un livello specifico"""
        return [n for n in self.nodes.values() if n.level == level]

    def get_bounds(self) -> Optional[Dict[str, float]]:
        """Calcola bounds delle coordinate"""
        if not self.nodes:
            return None

        nodes = list(self.nodes.values())
        return {
            'x_min': min(n.x for n in nodes),
            'x_max': max(n.x for n in nodes),
            'y_min': min(n.y for n in nodes),
            'y_max': max(n.y for n in nodes),
            'z_min': min(n.z for n in nodes),
            'z_max': max(n.z for n in nodes)
        }

    def clear(self):
        """Pulisce tutti i dati"""
        self.nodes.clear()
        self.edges.clear()
        self.annotations.clear()
        self.next_node_id = 1
        self.next_edge_id = 1
        self._mark_modified()

    def _mark_modified(self):
        """Marca come modificato"""
        self.metadata["modified"] = datetime.now().isoformat()

    # === Serializzazione ===

    def to_dict(self) -> Dict[str, Any]:
        """Esporta per JSON"""
        return {
            "version": "6.0",
            "metadata": self.metadata,
            "nodes": [node.to_dict() for node in self.nodes.values()],
            "edges": [edge.to_dict() for edge in self.edges.values()],
            "annotations": [ann.to_dict() for ann in self.annotations],
            "next_node_id": self.next_node_id,
            "next_edge_id": self.next_edge_id
        }

    def from_dict(self, data: Dict[str, Any]):
        """Importa da JSON"""
        self.clear()

        if "metadata" in data:
            self.metadata.update(data["metadata"])

        if "nodes" in data:
            for node_data in data["nodes"]:
                node = Node.from_dict(node_data)
                self.nodes[node.id] = node

        if "edges" in data:
            for edge_data in data["edges"]:
                edge = Edge.from_dict(edge_data)
                self.edges[edge.id] = edge

        if "annotations" in data:
            for ann_data in data["annotations"]:
                ann_type = ann_data.get('type')
                if ann_type == 'line':
                    self.annotations.append(LineAnnotation.from_dict(ann_data))
                elif ann_type == 'dimension':
                    self.annotations.append(DimensionAnnotation.from_dict(ann_data))

        self.next_node_id = data.get("next_node_id", 1)
        self.next_edge_id = data.get("next_edge_id", 1)

        # Correggi ID se necessario
        if self.nodes:
            max_id = max(self.nodes.keys())
            if max_id >= self.next_node_id:
                self.next_node_id = max_id + 1

        if self.edges:
            max_id = max(self.edges.keys())
            if max_id >= self.next_edge_id:
                self.next_edge_id = max_id + 1

    # === Undo/Redo ===

    def execute_command(self, command: 'Command') -> Any:
        """Esegue comando e aggiunge a undo stack"""
        result = command.execute(self)
        self.undo_stack.append(command)
        if len(self.undo_stack) > self.max_undo_levels:
            self.undo_stack.pop(0)
        self.redo_stack.clear()
        return result

    def undo(self) -> bool:
        """Annulla ultima operazione"""
        if not self.undo_stack:
            return False
        command = self.undo_stack.pop()
        command.undo(self)
        self.redo_stack.append(command)
        self._mark_modified()
        return True

    def redo(self) -> bool:
        """Ripristina operazione annullata"""
        if not self.redo_stack:
            return False
        command = self.redo_stack.pop()
        command.execute(self)
        self.undo_stack.append(command)
        self._mark_modified()
        return True

    def can_undo(self) -> bool:
        return len(self.undo_stack) > 0

    def can_redo(self) -> bool:
        return len(self.redo_stack) > 0

    def clear_undo_history(self):
        self.undo_stack.clear()
        self.redo_stack.clear()


# ============================================================================
# COMMAND PATTERN UNIFICATO
# ============================================================================

class Command:
    """Classe base per comandi Undo/Redo"""

    def execute(self, model: UnifiedModel) -> Any:
        """Esegue il comando"""
        raise NotImplementedError

    def undo(self, model: UnifiedModel):
        """Annulla il comando"""
        raise NotImplementedError


class CreateNodeCommand(Command):
    """Comando per creare un nodo"""

    def __init__(self, x: float, y: float, z: float = 0.0,
                 description: str = "", level: str = "",
                 topology_role: TopologyRole = TopologyRole.FREE):
        self.x = x
        self.y = y
        self.z = z
        self.description = description
        self.level = level
        self.topology_role = topology_role
        self.node_id = None

    def execute(self, model: UnifiedModel) -> int:
        self.node_id = model.add_node(self.x, self.y, self.z,
                                     self.description, self.level,
                                     self.topology_role)
        return self.node_id

    def undo(self, model: UnifiedModel):
        if self.node_id:
            model.delete_node(self.node_id)


class UpdateNodeCommand(Command):
    """Comando per aggiornare un nodo"""

    def __init__(self, node_id: int, **kwargs):
        self.node_id = node_id
        self.new_values = kwargs
        self.old_values = {}

    def execute(self, model: UnifiedModel) -> bool:
        if self.node_id in model.nodes:
            node = model.nodes[self.node_id]
            # Salva valori vecchi
            for key in self.new_values.keys():
                if hasattr(node, key):
                    self.old_values[key] = getattr(node, key)
            # Applica nuovi valori
            model.update_node(self.node_id, **self.new_values)
            return True
        return False

    def undo(self, model: UnifiedModel):
        if self.old_values:
            model.update_node(self.node_id, **self.old_values)


class DeleteNodeCommand(Command):
    """Comando per eliminare un nodo"""

    def __init__(self, node_id: int):
        self.node_id = node_id
        self.saved_node = None
        self.saved_edges = []

    def execute(self, model: UnifiedModel) -> bool:
        if self.node_id in model.nodes:
            # Salva nodo
            node = model.nodes[self.node_id]
            self.saved_node = node.to_dict()

            # Salva edges connessi
            for eid, edge in list(model.edges.items()):
                if edge.node1_id == self.node_id or edge.node2_id == self.node_id:
                    self.saved_edges.append(edge.to_dict())

            # Elimina
            model.delete_node(self.node_id)
            return True
        return False

    def undo(self, model: UnifiedModel):
        if self.saved_node:
            # Ripristina nodo
            node = Node.from_dict(self.saved_node)
            model.nodes[node.id] = node
            if node.id >= model.next_node_id:
                model.next_node_id = node.id + 1

            # Ripristina edges
            for edge_data in self.saved_edges:
                edge = Edge.from_dict(edge_data)
                model.edges[edge.id] = edge
                if edge.id >= model.next_edge_id:
                    model.next_edge_id = edge.id + 1


class CreateEdgeCommand(Command):
    """Comando per creare un edge"""

    def __init__(self, node1_id: int, node2_id: int,
                 edge_type: EdgeType = EdgeType.FREE):
        self.node1_id = node1_id
        self.node2_id = node2_id
        self.edge_type = edge_type
        self.edge_id = None

    def execute(self, model: UnifiedModel) -> Optional[int]:
        self.edge_id = model.add_edge(self.node1_id, self.node2_id, self.edge_type)
        return self.edge_id

    def undo(self, model: UnifiedModel):
        if self.edge_id:
            model.delete_edge(self.edge_id)


class DeleteEdgeCommand(Command):
    """Comando per eliminare un edge"""

    def __init__(self, edge_id: int):
        self.edge_id = edge_id
        self.saved_edge = None

    def execute(self, model: UnifiedModel) -> bool:
        if self.edge_id in model.edges:
            self.saved_edge = model.edges[self.edge_id].to_dict()
            model.delete_edge(self.edge_id)
            return True
        return False

    def undo(self, model: UnifiedModel):
        if self.saved_edge:
            edge = Edge.from_dict(self.saved_edge)
            model.edges[edge.id] = edge
            if edge.id >= model.next_edge_id:
                model.next_edge_id = edge.id + 1


# ============================================================================
# CANVAS VIEW COMPONENTS (semplificato per UnifiedModel)
# ============================================================================

class WireCanvasView(QWidget):
    """Canvas grafico per visualizzazione e editing nodi/edges"""

    selection_changed = pyqtSignal(int, str)  # node_id, type ('node'|'edge')

    def __init__(self, model: UnifiedModel, parent=None):
        super().__init__(parent)
        self.model = model
        self.setMinimumSize(800, 600)
        self.setMouseTracking(True)
        self.setFocusPolicy(Qt.StrongFocus)

        # Vista e navigazione
        self.zoom_factor = 1.0
        self.pan_offset = QPointF(0, 0)
        self.zoom_min = 0.1
        self.zoom_max = 10.0

        # Grid
        self.grid_size = 50  # pixel per metro base
        self.show_grid = True
        self.show_coordinates = True

        # Selezione
        self.selected_node_id = None
        self.selected_edge_id = None
        self.hover_node_id = None

        # Interazione
        self.dragging_node = False
        self.panning = False
        self.creating_edge = False
        self.edge_start_node_id = None
        self.drag_offset = QPointF(0, 0)
        self.last_mouse_pos = QPointF(0, 0)

        # Colori
        self.node_color = QColor(100, 150, 200)
        self.edge_color = QColor(80, 80, 80)
        self.selected_color = QColor(255, 100, 100)
        self.hover_color = QColor(100, 255, 100)
        self.grid_color = QColor(220, 220, 220)

        self.node_radius = 8

    def screen_to_world(self, screen_point: QPointF) -> QPointF:
        """Converte coordinate schermo in coordinate mondo (metri)"""
        adjusted = (screen_point - self.pan_offset) / self.zoom_factor
        x = adjusted.x() / self.grid_size
        y = (self.height() - adjusted.y()) / self.grid_size
        return QPointF(x, y)

    def world_to_screen(self, world_point: QPointF) -> QPointF:
        """Converte coordinate mondo in coordinate schermo"""
        x = world_point.x() * self.grid_size
        y = self.height() - (world_point.y() * self.grid_size)
        return QPointF(x, y) * self.zoom_factor + self.pan_offset

    def get_node_at(self, screen_point: QPointF, tolerance: float = 15) -> Optional[int]:
        """Trova nodo vicino al punto"""
        for node in self.model.nodes.values():
            world_pos = QPointF(node.x, node.y)
            screen_pos = self.world_to_screen(world_pos)
            dist = math.sqrt((screen_pos.x() - screen_point.x())**2 +
                           (screen_pos.y() - screen_point.y())**2)
            if dist <= tolerance:
                return node.id
        return None

    def paintEvent(self, event):
        """Rendering canvas"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Sfondo
        painter.fillRect(self.rect(), Qt.white)

        # Griglia
        if self.show_grid:
            self.draw_grid(painter)

        # Edges
        self.draw_edges(painter)

        # Nodes
        self.draw_nodes(painter)

        # Edge in creazione
        if self.creating_edge and self.edge_start_node_id:
            self.draw_temp_edge(painter)

    def draw_grid(self, painter: QPainter):
        """Disegna griglia"""
        painter.setPen(QPen(self.grid_color, 1))

        grid_step = self.grid_size * self.zoom_factor

        # Limita numero linee per performance
        if grid_step < 5:
            return

        # Linee verticali
        x = int(self.pan_offset.x() % grid_step)
        while x < self.width():
            painter.drawLine(int(x), 0, int(x), self.height())
            x += grid_step

        # Linee orizzontali
        y = int(self.pan_offset.y() % grid_step)
        while y < self.height():
            painter.drawLine(0, int(y), self.width(), int(y))
            y += grid_step

    def draw_edges(self, painter: QPainter):
        """Disegna edges"""
        for edge in self.model.edges.values():
            n1 = self.model.nodes.get(edge.node1_id)
            n2 = self.model.nodes.get(edge.node2_id)

            if not n1 or not n2:
                continue

            p1 = self.world_to_screen(QPointF(n1.x, n1.y))
            p2 = self.world_to_screen(QPointF(n2.x, n2.y))

            # Colore edge
            color = self.selected_color if edge.id == self.selected_edge_id else edge.color
            painter.setPen(QPen(color, edge.width))
            painter.drawLine(p1, p2)

            # Lunghezza edge
            if self.show_coordinates:
                length = n1.distance_to(n2)
                mid = QPointF((p1.x() + p2.x()) / 2, (p1.y() + p2.y()) / 2)
                painter.setPen(QPen(Qt.black))
                painter.setFont(QFont("Arial", 8))
                painter.drawText(mid, f"{length:.2f}m")

    def draw_nodes(self, painter: QPainter):
        """Disegna nodi"""
        for node in self.model.nodes.values():
            screen_pos = self.world_to_screen(QPointF(node.x, node.y))

            # Colore nodo
            if node.id == self.selected_node_id:
                color = self.selected_color
            elif node.id == self.hover_node_id:
                color = self.hover_color
            else:
                color = node.color

            # Disegna nodo
            painter.setBrush(QBrush(color))
            painter.setPen(QPen(Qt.black, 2))
            painter.drawEllipse(screen_pos, self.node_radius, self.node_radius)

            # ID e descrizione
            if self.show_coordinates:
                painter.setPen(QPen(Qt.black))
                painter.setFont(QFont("Arial", 9))
                text = f"#{node.id}"
                if node.description:
                    text += f" {node.description}"
                painter.drawText(screen_pos + QPointF(self.node_radius + 5, 5), text)

                # Coordinate
                painter.setFont(QFont("Arial", 8))
                coord_text = f"({node.x:.2f}, {node.y:.2f})"
                painter.drawText(screen_pos + QPointF(self.node_radius + 5, 18), coord_text)

    def draw_temp_edge(self, painter: QPainter):
        """Disegna edge temporaneo durante creazione"""
        if not self.edge_start_node_id:
            return

        start_node = self.model.nodes.get(self.edge_start_node_id)
        if not start_node:
            return

        p1 = self.world_to_screen(QPointF(start_node.x, start_node.y))
        p2 = self.last_mouse_pos

        painter.setPen(QPen(QColor(255, 0, 0), 2, Qt.DashLine))
        painter.drawLine(p1, p2)

    def mousePressEvent(self, event):
        """Mouse press - selezione, drag, pan"""
        pos = QPointF(event.pos())
        self.last_mouse_pos = pos

        if event.button() == Qt.LeftButton:
            # Cerca nodo
            node_id = self.get_node_at(pos)

            if node_id:
                if event.modifiers() & Qt.ControlModifier:
                    # Ctrl+Click = inizia edge
                    self.creating_edge = True
                    self.edge_start_node_id = node_id
                else:
                    # Selezione e drag
                    self.selected_node_id = node_id
                    self.selected_edge_id = None
                    self.dragging_node = True

                    node = self.model.nodes[node_id]
                    screen_pos = self.world_to_screen(QPointF(node.x, node.y))
                    self.drag_offset = screen_pos - pos

                    self.selection_changed.emit(node_id, 'node')
            else:
                # Click su vuoto = deseleziona
                self.selected_node_id = None
                self.selected_edge_id = None
                self.selection_changed.emit(-1, 'none')

        elif event.button() == Qt.MiddleButton or event.button() == Qt.RightButton:
            # Pan
            self.panning = True
            self.setCursor(Qt.ClosedHandCursor)

        self.update()

    def mouseMoveEvent(self, event):
        """Mouse move - drag, pan, hover"""
        pos = QPointF(event.pos())
        delta = pos - self.last_mouse_pos
        self.last_mouse_pos = pos

        if self.dragging_node and self.selected_node_id:
            # Drag nodo
            new_screen_pos = pos + self.drag_offset
            new_world_pos = self.screen_to_world(new_screen_pos)

            # Aggiorna nodo con Command
            cmd = UpdateNodeCommand(self.selected_node_id, x=new_world_pos.x(), y=new_world_pos.y())
            self.model.execute_command(cmd)

        elif self.panning:
            # Pan vista
            self.pan_offset += delta

        elif self.creating_edge:
            # Aggiorna preview edge
            pass

        else:
            # Hover detection
            hover_id = self.get_node_at(pos)
            if hover_id != self.hover_node_id:
                self.hover_node_id = hover_id
                self.update()
                return

        self.update()

    def mouseReleaseEvent(self, event):
        """Mouse release - fine drag/pan/edge"""
        if event.button() == Qt.LeftButton:
            if self.creating_edge and self.edge_start_node_id:
                # Fine creazione edge
                pos = QPointF(event.pos())
                end_node_id = self.get_node_at(pos)

                if end_node_id and end_node_id != self.edge_start_node_id:
                    # Crea edge
                    cmd = CreateEdgeCommand(self.edge_start_node_id, end_node_id, EdgeType.FREE)
                    self.model.execute_command(cmd)

                self.creating_edge = False
                self.edge_start_node_id = None

            self.dragging_node = False

        elif event.button() == Qt.MiddleButton or event.button() == Qt.RightButton:
            self.panning = False
            self.setCursor(Qt.ArrowCursor)

        self.update()

    def wheelEvent(self, event):
        """Wheel - zoom"""
        zoom_delta = event.angleDelta().y()

        if zoom_delta > 0:
            self.zoom_factor *= 1.1
        else:
            self.zoom_factor *= 0.9

        self.zoom_factor = max(self.zoom_min, min(self.zoom_max, self.zoom_factor))
        self.update()

    def keyPressEvent(self, event):
        """Gestione tasti"""
        if event.key() == Qt.Key_Delete and self.selected_node_id:
            # Elimina nodo selezionato
            reply = QMessageBox.question(
                self, 'Conferma',
                f'Eliminare nodo ID {self.selected_node_id}?',
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                cmd = DeleteNodeCommand(self.selected_node_id)
                self.model.execute_command(cmd)
                self.selected_node_id = None
                self.update()

        elif event.key() == Qt.Key_F:
            # Fit to extents
            self.fit_to_extents()

    def mouseDoubleClickEvent(self, event):
        """Double click - crea nodo"""
        if event.button() == Qt.LeftButton:
            pos = QPointF(event.pos())
            world_pos = self.screen_to_world(pos)

            # Crea nuovo nodo
            cmd = CreateNodeCommand(world_pos.x(), world_pos.y(), 0.0, "", "", TopologyRole.FREE)
            node_id = self.model.execute_command(cmd)
            self.selected_node_id = node_id
            self.selection_changed.emit(node_id, 'node')
            self.update()

    def fit_to_extents(self):
        """Adatta vista a tutti i nodi"""
        if not self.model.nodes:
            return

        bounds = self.model.get_bounds()
        if not bounds:
            return

        margin = 1.0  # metri
        x_min = bounds['x_min'] - margin
        x_max = bounds['x_max'] + margin
        y_min = bounds['y_min'] - margin
        y_max = bounds['y_max'] + margin

        width = x_max - x_min
        height = y_max - y_min

        if width <= 0 or height <= 0:
            return

        # Calcola zoom per contenere tutto
        zoom_x = self.width() / (width * self.grid_size)
        zoom_y = self.height() / (height * self.grid_size)
        self.zoom_factor = min(zoom_x, zoom_y) * 0.9  # 90% per margine

        self.zoom_factor = max(self.zoom_min, min(self.zoom_max, self.zoom_factor))

        # Centra
        center_x = (x_min + x_max) / 2
        center_y = (y_min + y_max) / 2
        self.center_on_world_point(QPointF(center_x, center_y))

    def center_on_world_point(self, world_point: QPointF):
        """Centra vista su punto mondo"""
        screen_center = QPointF(self.width() / 2, self.height() / 2)
        target_screen = self.world_to_screen(world_point)
        self.pan_offset += screen_center - target_screen
        self.update()

    def refresh(self):
        """Aggiorna canvas"""
        self.update()


# ============================================================================
# 3D VISUALIZATION COMPONENTS
# ============================================================================

class Visualization3DWidget(QWidget):
    """Widget visualizzazione 3D con matplotlib"""

    def __init__(self, model: UnifiedModel, parent=None):
        super().__init__(parent)
        self.model = model
        self.setup_ui()

    def setup_ui(self):
        """Setup interfaccia"""
        layout = QVBoxLayout()

        if not MATPLOTLIB_AVAILABLE:
            label = QLabel("⚠️ Matplotlib non disponibile\n\nInstallare con: pip install matplotlib")
            label.setAlignment(Qt.AlignCenter)
            label.setFont(QFont("Arial", 12))
            layout.addWidget(label)
            self.setLayout(layout)
            return

        # Canvas matplotlib
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar

        self.figure = Figure(figsize=(10, 8))
        self.canvas = FigureCanvas(self.figure)
        self.ax = self.figure.add_subplot(111, projection='3d')

        # Toolbar matplotlib
        self.toolbar = NavigationToolbar(self.canvas, self)
        layout.addWidget(self.toolbar)
        layout.addWidget(self.canvas)

        # Controlli vista
        controls_layout = QHBoxLayout()

        # Vista predefinite
        btn_xy = QPushButton("Vista XY (Top)")
        btn_xy.clicked.connect(lambda: self.set_view(90, -90))
        controls_layout.addWidget(btn_xy)

        btn_xz = QPushButton("Vista XZ (Front)")
        btn_xz.clicked.connect(lambda: self.set_view(0, -90))
        controls_layout.addWidget(btn_xz)

        btn_yz = QPushButton("Vista YZ (Side)")
        btn_yz.clicked.connect(lambda: self.set_view(0, 0))
        controls_layout.addWidget(btn_yz)

        btn_iso = QPushButton("Vista Isometrica")
        btn_iso.clicked.connect(lambda: self.set_view(30, -45))
        controls_layout.addWidget(btn_iso)

        controls_layout.addStretch()

        # Checkbox visualizzazione
        self.chk_nodes = QCheckBox("Nodi")
        self.chk_nodes.setChecked(True)
        self.chk_nodes.toggled.connect(self.refresh_plot)
        controls_layout.addWidget(self.chk_nodes)

        self.chk_edges = QCheckBox("Edges")
        self.chk_edges.setChecked(True)
        self.chk_edges.toggled.connect(self.refresh_plot)
        controls_layout.addWidget(self.chk_edges)

        self.chk_labels = QCheckBox("Labels")
        self.chk_labels.setChecked(True)
        self.chk_labels.toggled.connect(self.refresh_plot)
        controls_layout.addWidget(self.chk_labels)

        btn_refresh = QPushButton("🔄 Refresh")
        btn_refresh.clicked.connect(self.refresh_plot)
        controls_layout.addWidget(btn_refresh)

        layout.addLayout(controls_layout)

        self.setLayout(layout)

        # Plot iniziale
        self.refresh_plot()

    def set_view(self, elev: float, azim: float):
        """Imposta vista 3D"""
        if MATPLOTLIB_AVAILABLE:
            self.ax.view_init(elev=elev, azim=azim)
            self.canvas.draw()

    def refresh_plot(self):
        """Aggiorna plot 3D"""
        if not MATPLOTLIB_AVAILABLE:
            return

        self.ax.clear()

        if not self.model.nodes:
            self.ax.text(0.5, 0.5, 0.5, 'Nessun nodo da visualizzare',
                        ha='center', va='center', fontsize=12)
            self.canvas.draw()
            return

        # Raccogli coordinate
        nodes = list(self.model.nodes.values())
        xs = [n.x for n in nodes]
        ys = [n.y for n in nodes]
        zs = [n.z for n in nodes]

        # Disegna edges
        if self.chk_edges.isChecked():
            for edge in self.model.edges.values():
                n1 = self.model.nodes.get(edge.node1_id)
                n2 = self.model.nodes.get(edge.node2_id)
                if n1 and n2:
                    self.ax.plot([n1.x, n2.x], [n1.y, n2.y], [n1.z, n2.z],
                                'k-', linewidth=1, alpha=0.6)

        # Disegna nodi
        if self.chk_nodes.isChecked():
            # Colora per livello
            levels = set(n.level for n in nodes if n.level)
            if levels:
                import matplotlib.cm as cm
                colors_map = cm.get_cmap('tab10')
                level_to_color = {level: colors_map(i / max(len(levels), 1))
                                 for i, level in enumerate(sorted(levels))}

                for level in levels:
                    level_nodes = [n for n in nodes if n.level == level]
                    if level_nodes:
                        lxs = [n.x for n in level_nodes]
                        lys = [n.y for n in level_nodes]
                        lzs = [n.z for n in level_nodes]
                        self.ax.scatter(lxs, lys, lzs, c=[level_to_color[level]],
                                      s=50, label=f"Livello {level}", alpha=0.8)

                # Nodi senza livello
                no_level = [n for n in nodes if not n.level]
                if no_level:
                    nlxs = [n.x for n in no_level]
                    nlys = [n.y for n in no_level]
                    nlzs = [n.z for n in no_level]
                    self.ax.scatter(nlxs, nlys, nlzs, c='gray',
                                  s=50, label="Senza livello", alpha=0.8)

                self.ax.legend()
            else:
                # Tutti dello stesso colore
                self.ax.scatter(xs, ys, zs, c='blue', s=50, alpha=0.8)

        # Labels
        if self.chk_labels.isChecked():
            for node in nodes:
                label = f"#{node.id}"
                if node.description:
                    label += f"\n{node.description}"
                self.ax.text(node.x, node.y, node.z, label,
                           fontsize=8, ha='left')

        # Assi e titolo
        self.ax.set_xlabel('X (m)', fontsize=10)
        self.ax.set_ylabel('Y (m)', fontsize=10)
        self.ax.set_zlabel('Z (m)', fontsize=10)
        self.ax.set_title(f'Visualizzazione 3D - {len(nodes)} nodi, {len(self.model.edges)} edges',
                         fontsize=12, fontweight='bold')

        # Griglia
        self.ax.grid(True, alpha=0.3)

        # Aspect ratio uguale
        if nodes:
            bounds = self.model.get_bounds()
            if bounds:
                # Set limits con margine
                margin = 1.0
                self.ax.set_xlim(bounds['x_min'] - margin, bounds['x_max'] + margin)
                self.ax.set_ylim(bounds['y_min'] - margin, bounds['y_max'] + margin)
                self.ax.set_zlim(bounds['z_min'] - margin, bounds['z_max'] + margin)

        self.canvas.draw()


# ============================================================================
# TABLE VIEW COMPONENTS (da wire_editor.py)
# ============================================================================

class CoordinateInputPanel(QGroupBox):
    """Pannello input coordinate per UnifiedModel"""

    node_added = pyqtSignal(int)  # Segnale quando aggiunto nodo

    def __init__(self, model: UnifiedModel, parent=None):
        super().__init__("Input Coordinate", parent)
        self.model = model
        self.setup_ui()

    def setup_ui(self):
        layout = QGridLayout()

        # Coordinate principali
        layout.addWidget(QLabel("X (m):"), 0, 0)
        self.x_spin = QDoubleSpinBox()
        self.x_spin.setRange(-10000, 10000)
        self.x_spin.setDecimals(4)
        self.x_spin.setSingleStep(0.1)
        self.x_spin.setMinimumWidth(120)
        layout.addWidget(self.x_spin, 0, 1)

        layout.addWidget(QLabel("Y (m):"), 0, 2)
        self.y_spin = QDoubleSpinBox()
        self.y_spin.setRange(-10000, 10000)
        self.y_spin.setDecimals(4)
        self.y_spin.setSingleStep(0.1)
        self.y_spin.setMinimumWidth(120)
        layout.addWidget(self.y_spin, 0, 3)

        layout.addWidget(QLabel("Z (m):"), 1, 0)
        self.z_spin = QDoubleSpinBox()
        self.z_spin.setRange(-1000, 1000)
        self.z_spin.setDecimals(3)
        self.z_spin.setSingleStep(0.1)
        self.z_spin.setMinimumWidth(120)
        layout.addWidget(self.z_spin, 1, 1)

        # Livello
        layout.addWidget(QLabel("Livello:"), 1, 2)
        self.level_edit = QLineEdit()
        self.level_edit.setPlaceholderText("es: PT, P1, P2")
        self.level_edit.setMaximumWidth(120)
        layout.addWidget(self.level_edit, 1, 3)

        # Descrizione
        layout.addWidget(QLabel("Descrizione:"), 2, 0)
        self.desc_edit = QLineEdit()
        self.desc_edit.setPlaceholderText("Opzionale - es: Pilastro A1")
        layout.addWidget(self.desc_edit, 2, 1, 1, 3)

        # Pulsanti
        button_layout = QHBoxLayout()

        self.add_btn = QPushButton("Aggiungi Nodo")
        self.add_btn.setDefault(True)
        self.add_btn.clicked.connect(self.add_node)
        button_layout.addWidget(self.add_btn)

        self.clear_btn = QPushButton("Reset")
        self.clear_btn.clicked.connect(self.clear_fields)
        button_layout.addWidget(self.clear_btn)

        layout.addLayout(button_layout, 3, 0, 1, 4)
        self.setLayout(layout)

        # Focus su X per input rapido
        self.x_spin.setFocus()

        # Enter per aggiungere
        self.desc_edit.returnPressed.connect(self.add_node)

    def add_node(self):
        """Aggiunge nodo usando Command pattern"""
        x = self.x_spin.value()
        y = self.y_spin.value()
        z = self.z_spin.value()
        description = self.desc_edit.text().strip()
        level = self.level_edit.text().strip()

        # Usa Command per undo/redo
        cmd = CreateNodeCommand(x, y, z, description, level, TopologyRole.FREE)
        node_id = self.model.execute_command(cmd)

        # Emetti segnale
        self.node_added.emit(node_id)

        # Incrementa automaticamente X per input sequenziale
        self.x_spin.setValue(x + 1.0)
        self.desc_edit.clear()
        self.x_spin.setFocus()

        return node_id

    def clear_fields(self):
        """Reset tutti i campi"""
        self.x_spin.setValue(0.0)
        self.y_spin.setValue(0.0)
        self.z_spin.setValue(0.0)
        self.level_edit.clear()
        self.desc_edit.clear()
        self.x_spin.setFocus()

    def keyPressEvent(self, event):
        """Enter = aggiungi nodo"""
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            self.add_node()
        super().keyPressEvent(event)


class CoordinateTable(QTableWidget):
    """Tabella coordinate con editing inline per UnifiedModel"""

    data_changed = pyqtSignal()  # Segnale quando dati cambiano

    def __init__(self, model: UnifiedModel, parent=None):
        super().__init__(parent)
        self.model = model
        self.setup_table()

    def setup_table(self):
        self.setColumnCount(6)
        self.setHorizontalHeaderLabels(['ID', 'X (m)', 'Y (m)', 'Z (m)', 'Livello', 'Descrizione'])

        # Configurazione header
        header = self.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID fisso
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # X
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Y
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Z
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Livello
        header.setStretchLastSection(True)  # Descrizione espandibile

        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(True)

        # Connetti modifiche
        self.cellChanged.connect(self.on_cell_changed)

    def refresh(self, filter_level: Optional[str] = None):
        """Aggiorna tabella completa con filtro opzionale per livello"""
        nodes = self.model.get_node_list()

        # Applica filtro livello se specificato
        if filter_level and filter_level != "Tutti":
            nodes = [n for n in nodes if n.level == filter_level]

        self.setRowCount(len(nodes))

        # Blocca segnali durante refresh
        self.blockSignals(True)

        for row, node in enumerate(nodes):
            # ID (solo lettura)
            id_item = QTableWidgetItem(str(node.id))
            id_item.setFlags(id_item.flags() & ~Qt.ItemIsEditable)
            self.setItem(row, 0, id_item)

            # Coordinate (modificabili con precisione)
            self.setItem(row, 1, QTableWidgetItem(f"{node.x:.4f}"))
            self.setItem(row, 2, QTableWidgetItem(f"{node.y:.4f}"))
            self.setItem(row, 3, QTableWidgetItem(f"{node.z:.3f}"))

            # Livello
            self.setItem(row, 4, QTableWidgetItem(node.level))

            # Descrizione
            self.setItem(row, 5, QTableWidgetItem(node.description))

        self.blockSignals(False)
        # Non emettere data_changed qui - causa ricorsione infinita
        # Il segnale viene emesso solo in on_cell_changed quando dati effettivamente cambiano

    def on_cell_changed(self, row: int, col: int):
        """Gestisce modifiche dirette in tabella usando Command pattern"""
        id_item = self.item(row, 0)
        if not id_item:
            return

        node_id = int(id_item.text())
        item = self.item(row, col)

        try:
            if col == 1:  # X
                cmd = UpdateNodeCommand(node_id, x=float(item.text()))
                self.model.execute_command(cmd)
            elif col == 2:  # Y
                cmd = UpdateNodeCommand(node_id, y=float(item.text()))
                self.model.execute_command(cmd)
            elif col == 3:  # Z
                cmd = UpdateNodeCommand(node_id, z=float(item.text()))
                self.model.execute_command(cmd)
            elif col == 4:  # Livello
                cmd = UpdateNodeCommand(node_id, level=item.text())
                self.model.execute_command(cmd)
            elif col == 5:  # Descrizione
                cmd = UpdateNodeCommand(node_id, description=item.text())
                self.model.execute_command(cmd)

            self.data_changed.emit()
        except ValueError as e:
            # Valore non valido, ripristina
            QMessageBox.warning(self, "Errore", f"Valore non valido: {e}")
            self.refresh()

    def keyPressEvent(self, event):
        """Elimina nodo con DEL"""
        if event.key() == Qt.Key_Delete:
            current_row = self.currentRow()
            if current_row >= 0:
                id_item = self.item(current_row, 0)
                if id_item:
                    node_id = int(id_item.text())
                    reply = QMessageBox.question(
                        self, 'Conferma Eliminazione',
                        f'Eliminare il nodo ID {node_id}?',
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if reply == QMessageBox.Yes:
                        cmd = DeleteNodeCommand(node_id)
                        self.model.execute_command(cmd)
                        self.refresh()
                        self.data_changed.emit()
        super().keyPressEvent(event)


# ============================================================================
# MAIN WINDOW UNIFICATA
# ============================================================================

class UnifiedMainWindow(QMainWindow):
    """Finestra principale unificata con tab per diverse viste"""

    def __init__(self):
        super().__init__()
        self.model = UnifiedModel()
        self.setWindowTitle("Wire Editor Unified v6.1 - Arch. Michelangelo Bartolotta")
        self.setMinimumSize(1400, 900)

        self.setup_ui()
        self.setup_menus()
        self.setup_toolbar()
        self.setup_statusbar()

    def setup_ui(self):
        """Setup interfaccia con tab"""
        # Central widget con tab
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.currentChanged.connect(self.on_tab_changed)

        # Tab 1: Table View (da wire_editor.py)
        self.table_tab = QWidget()
        self.setup_table_tab()
        self.tabs.addTab(self.table_tab, "📊 Table View")

        # Tab 2: Canvas View (da wire_editor1.py)
        self.canvas_tab = QWidget()
        self.setup_canvas_tab()
        self.tabs.addTab(self.canvas_tab, "🎨 Canvas View")

        # Tab 3: 3D Visualization
        self.viz_tab = QWidget()
        self.setup_viz_tab()
        self.tabs.addTab(self.viz_tab, "📈 3D View")

        # Tab 4: Reports
        self.reports_tab = QWidget()
        self.setup_reports_tab()
        self.tabs.addTab(self.reports_tab, "📄 Reports")

        self.setCentralWidget(self.tabs)

    def setup_table_tab(self):
        """Setup tab vista tabellare (integrato da wire_editor.py)"""
        layout = QVBoxLayout()

        # Pannello input coordinate
        self.input_panel = CoordinateInputPanel(self.model)
        self.input_panel.node_added.connect(self.on_node_added)
        layout.addWidget(self.input_panel)

        # Tabella coordinate con filtro
        table_group = QGroupBox("Elenco Nodi")
        table_layout = QVBoxLayout()

        self.coord_table = CoordinateTable(self.model)
        self.coord_table.data_changed.connect(self.on_table_data_changed)
        table_layout.addWidget(self.coord_table)

        table_group.setLayout(table_layout)
        layout.addWidget(table_group)

        # Barra informazioni e filtri
        info_layout = QHBoxLayout()

        self.table_info_label = QLabel("Nodi: 0")
        self.table_info_label.setFont(QFont("Arial", 10, QFont.Bold))
        info_layout.addWidget(self.table_info_label)

        info_layout.addStretch()

        # Filtro livello
        info_layout.addWidget(QLabel("Filtro Livello:"))
        self.level_filter = QComboBox()
        self.level_filter.addItem("Tutti")
        self.level_filter.currentTextChanged.connect(self.on_level_filter_changed)
        self.level_filter.setMinimumWidth(120)
        info_layout.addWidget(self.level_filter)

        info_layout.addStretch()

        self.bounds_label = QLabel("")
        self.bounds_label.setFont(QFont("Arial", 9))
        info_layout.addWidget(self.bounds_label)

        layout.addLayout(info_layout)

        self.table_tab.setLayout(layout)

        # Refresh iniziale
        self.refresh_table_view()

    def setup_canvas_tab(self):
        """Setup tab canvas grafico (integrato)"""
        layout = QVBoxLayout()

        # Canvas
        self.canvas_view = WireCanvasView(self.model)
        self.canvas_view.selection_changed.connect(self.on_canvas_selection_changed)
        layout.addWidget(self.canvas_view)

        # Toolbar canvas
        toolbar_layout = QHBoxLayout()

        # Pulsanti azione
        btn_fit = QPushButton("Fit All (F)")
        btn_fit.clicked.connect(self.canvas_view.fit_to_extents)
        toolbar_layout.addWidget(btn_fit)

        # Toggle griglia
        self.chk_grid = QCheckBox("Griglia")
        self.chk_grid.setChecked(self.canvas_view.show_grid)
        self.chk_grid.toggled.connect(self.on_grid_toggled)
        toolbar_layout.addWidget(self.chk_grid)

        # Toggle coordinate
        self.chk_coords = QCheckBox("Coordinate")
        self.chk_coords.setChecked(self.canvas_view.show_coordinates)
        self.chk_coords.toggled.connect(self.on_coords_toggled)
        toolbar_layout.addWidget(self.chk_coords)

        toolbar_layout.addStretch()

        # Info selezione
        self.canvas_info_label = QLabel("Nessuna selezione")
        self.canvas_info_label.setFont(QFont("Arial", 9))
        toolbar_layout.addWidget(self.canvas_info_label)

        layout.addLayout(toolbar_layout)

        # Istruzioni
        help_text = QLabel("💡 Doppio click: crea nodo | Drag: sposta | Ctrl+Click: crea edge | Tasto centrale/destro: pan | Rotella: zoom")
        help_text.setFont(QFont("Arial", 8))
        help_text.setStyleSheet("color: #666; padding: 5px;")
        layout.addWidget(help_text)

        self.canvas_tab.setLayout(layout)

    def setup_viz_tab(self):
        """Setup tab visualizzazione 3D (integrato)"""
        layout = QVBoxLayout()

        # Widget 3D
        self.viz_3d = Visualization3DWidget(self.model)
        layout.addWidget(self.viz_3d)

        self.viz_tab.setLayout(layout)

    def setup_reports_tab(self):
        """Setup tab reports (integrato)"""
        layout = QVBoxLayout()

        # Titolo
        title = QLabel("Reports & Export")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)

        # Import section
        import_group = QGroupBox("Import Dati")
        import_layout = QVBoxLayout()

        # Pulsanti import
        import_btn_layout = QHBoxLayout()

        btn_import_csv = QPushButton("📄 Import CSV")
        btn_import_csv.clicked.connect(self.import_csv)
        import_btn_layout.addWidget(btn_import_csv)

        btn_import_excel = QPushButton("📊 Import Excel")
        btn_import_excel.clicked.connect(self.import_excel)
        if not OPENPYXL_AVAILABLE:
            btn_import_excel.setEnabled(False)
            btn_import_excel.setToolTip("Openpyxl non installato - pip install openpyxl")
        import_btn_layout.addWidget(btn_import_excel)

        btn_import_dxf = QPushButton("📐 Import DXF")
        btn_import_dxf.clicked.connect(self.import_dxf)
        import_btn_layout.addWidget(btn_import_dxf)

        import_btn_layout.addStretch()

        import_layout.addLayout(import_btn_layout)
        import_group.setLayout(import_layout)
        layout.addWidget(import_group)

        # Export section
        export_group = QGroupBox("Export Dati")
        export_layout = QVBoxLayout()

        # Pulsanti export
        btn_layout = QHBoxLayout()

        btn_csv = QPushButton("📄 Export CSV")
        btn_csv.clicked.connect(self.export_csv)
        btn_layout.addWidget(btn_csv)

        btn_excel = QPushButton("📊 Export Excel")
        btn_excel.clicked.connect(self.export_excel)
        if not OPENPYXL_AVAILABLE:
            btn_excel.setEnabled(False)
            btn_excel.setToolTip("Openpyxl non installato - pip install openpyxl")
        btn_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("📑 Export PDF Report")
        btn_pdf.clicked.connect(self.export_pdf_report)
        if not REPORTLAB_AVAILABLE:
            btn_pdf.setEnabled(False)
            btn_pdf.setToolTip("Reportlab non installato - pip install reportlab")
        btn_layout.addWidget(btn_pdf)

        btn_layout.addStretch()

        export_layout.addLayout(btn_layout)
        export_group.setLayout(export_layout)
        layout.addWidget(export_group)

        # Statistiche section
        stats_group = QGroupBox("Statistiche Progetto")
        stats_layout = QVBoxLayout()

        self.stats_text = QTextEdit()
        self.stats_text.setReadOnly(True)
        self.stats_text.setMaximumHeight(300)
        stats_layout.addWidget(self.stats_text)

        btn_refresh_stats = QPushButton("🔄 Aggiorna Statistiche")
        btn_refresh_stats.clicked.connect(self.update_statistics)
        stats_layout.addWidget(btn_refresh_stats)

        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)

        layout.addStretch()

        self.reports_tab.setLayout(layout)

        # Statistiche iniziali
        self.update_statistics()

    def setup_menus(self):
        """Setup menu bar unificato"""
        menubar = self.menuBar()

        # File menu
        file_menu = menubar.addMenu('File')
        file_menu.addAction('New Project', self.new_project, 'Ctrl+N')
        file_menu.addSeparator()
        file_menu.addAction('Open...', self.open_file, 'Ctrl+O')
        file_menu.addAction('Save...', self.save_file, 'Ctrl+S')
        file_menu.addSeparator()

        # Import submenu
        import_menu = file_menu.addMenu('Import')
        import_menu.addAction('Import CSV...', self.import_csv)
        import_menu.addAction('Import Excel...', self.import_excel)
        import_menu.addAction('Import DXF...', self.import_dxf)

        file_menu.addSeparator()
        file_menu.addAction('Exit', self.close, 'Ctrl+Q')

        # Edit menu
        edit_menu = menubar.addMenu('Edit')
        self.undo_action = edit_menu.addAction('Undo', self.undo, 'Ctrl+Z')
        self.redo_action = edit_menu.addAction('Redo', self.redo, 'Ctrl+Y')
        edit_menu.addSeparator()
        edit_menu.addAction('Project Info...', self.edit_metadata)

        # View menu
        view_menu = menubar.addMenu('View')
        view_menu.addAction('Table View', lambda: self.tabs.setCurrentIndex(0), 'Ctrl+1')
        view_menu.addAction('Canvas View', lambda: self.tabs.setCurrentIndex(1), 'Ctrl+2')
        view_menu.addAction('3D View', lambda: self.tabs.setCurrentIndex(2), 'Ctrl+3')
        view_menu.addAction('Reports', lambda: self.tabs.setCurrentIndex(3), 'Ctrl+4')

        # Help menu
        help_menu = menubar.addMenu('Help')
        help_menu.addAction('About', self.show_about)

        # Aggiorna stato undo/redo
        self.update_undo_redo_actions()

    def setup_toolbar(self):
        """Setup toolbar"""
        toolbar = QToolBar("Main Toolbar")
        self.addToolBar(toolbar)

        # Actions base
        toolbar.addAction("New", self.new_project)
        toolbar.addAction("Open", self.open_file)
        toolbar.addAction("Save", self.save_file)
        toolbar.addSeparator()
        toolbar.addAction("Undo", self.undo)
        toolbar.addAction("Redo", self.redo)

    def setup_statusbar(self):
        """Setup status bar"""
        self.status_label = QLabel("Ready - Wire Editor Unified v6.1")
        self.statusBar().addWidget(self.status_label)

        self.node_count_label = QLabel("Nodes: 0")
        self.statusBar().addPermanentWidget(self.node_count_label)

        self.edge_count_label = QLabel("Edges: 0")
        self.statusBar().addPermanentWidget(self.edge_count_label)

    def update_status(self):
        """Aggiorna status bar"""
        self.node_count_label.setText(f"Nodes: {len(self.model.nodes)}")
        self.edge_count_label.setText(f"Edges: {len(self.model.edges)}")

    def update_undo_redo_actions(self):
        """Aggiorna stato azioni undo/redo"""
        self.undo_action.setEnabled(self.model.can_undo())
        self.redo_action.setEnabled(self.model.can_redo())

        undo_count = len(self.model.undo_stack)
        redo_count = len(self.model.redo_stack)
        self.undo_action.setText(f"Undo ({undo_count})" if undo_count > 0 else "Undo")
        self.redo_action.setText(f"Redo ({redo_count})" if redo_count > 0 else "Redo")

    # === Slot menu ===

    def new_project(self):
        """Nuovo progetto"""
        if len(self.model.nodes) > 0 or len(self.model.edges) > 0:
            reply = QMessageBox.question(
                self, 'New Project',
                'Clear all data and start new project?',
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.No:
                return

        self.model.clear()
        self.model.clear_undo_history()
        self.update_status()
        self.update_undo_redo_actions()
        QMessageBox.information(self, "New Project", "Project initialized")

    def open_file(self):
        """Apri file progetto"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "Open Project", "",
            "JSON Files (*.json);;All Files (*)"
        )

        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.model.from_dict(data)
                self.model.clear_undo_history()
                self.update_status()
                self.update_undo_redo_actions()
                QMessageBox.information(self, "Open Project",
                                      f"Project loaded from:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Error",
                                   f"Cannot open file:\n{e}")

    def save_file(self):
        """Salva file progetto"""
        if not self.model.nodes and not self.model.edges:
            QMessageBox.warning(self, "Warning", "No data to save")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Save Project", "project.json",
            "JSON Files (*.json);;All Files (*)"
        )

        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(self.model.to_dict(), f, indent=2, ensure_ascii=False)
                QMessageBox.information(self, "Save Project",
                                      f"Project saved to:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Error",
                                   f"Cannot save file:\n{e}")

    def undo(self):
        """Annulla ultima operazione"""
        if self.model.undo():
            self.update_status()
            self.update_undo_redo_actions()
        else:
            QMessageBox.information(self, "Undo", "Nothing to undo")

    def redo(self):
        """Ripristina operazione"""
        if self.model.redo():
            self.update_status()
            self.update_undo_redo_actions()
        else:
            QMessageBox.information(self, "Redo", "Nothing to redo")

    def edit_metadata(self):
        """Modifica metadata progetto"""
        # TODO: Implementare dialog metadata
        QMessageBox.information(self, "Project Info", "FASE 2: Metadata editor")

    def on_tab_changed(self, index: int):
        """Gestisce cambio tab"""
        tab_names = ["Table View", "Canvas View", "3D View", "Reports"]
        if index < len(tab_names):
            self.status_label.setText(f"Active: {tab_names[index]}")

    # === Callbacks Table View ===

    def on_node_added(self, node_id: int):
        """Callback quando aggiunto nodo"""
        self.refresh_table_view()
        self.canvas_view.refresh()
        if MATPLOTLIB_AVAILABLE and hasattr(self, 'viz_3d'):
            self.viz_3d.refresh_plot()
        self.update_status()
        self.update_undo_redo_actions()

    def on_table_data_changed(self):
        """Callback quando dati tabella cambiano"""
        self.refresh_table_view()
        self.canvas_view.refresh()
        if MATPLOTLIB_AVAILABLE and hasattr(self, 'viz_3d'):
            self.viz_3d.refresh_plot()
        self.update_status()
        self.update_undo_redo_actions()

    def on_level_filter_changed(self, level: str):
        """Gestisce cambio filtro livello"""
        self.coord_table.refresh(level if level != "Tutti" else None)
        self.update_table_info()

    # === Callbacks Canvas View ===

    def on_canvas_selection_changed(self, node_id: int, sel_type: str):
        """Callback quando cambia selezione nel canvas"""
        if sel_type == 'node' and node_id >= 0:
            node = self.model.nodes.get(node_id)
            if node:
                self.canvas_info_label.setText(
                    f"Nodo #{node.id}: ({node.x:.2f}, {node.y:.2f}) - {node.description}"
                )
        else:
            self.canvas_info_label.setText("Nessuna selezione")

        # Aggiorna anche tabella e 3D se modifiche dal canvas
        self.refresh_table_view()
        if MATPLOTLIB_AVAILABLE and hasattr(self, 'viz_3d'):
            self.viz_3d.refresh_plot()
        self.update_status()
        self.update_undo_redo_actions()

    def on_grid_toggled(self, checked: bool):
        """Toggle griglia canvas"""
        self.canvas_view.show_grid = checked
        self.canvas_view.update()

    def on_coords_toggled(self, checked: bool):
        """Toggle coordinate canvas"""
        self.canvas_view.show_coordinates = checked
        self.canvas_view.update()

    def refresh_table_view(self):
        """Aggiorna vista tabellare completa"""
        current_filter = self.level_filter.currentText()
        self.coord_table.refresh(current_filter if current_filter != "Tutti" else None)

        # Aggiorna combo livelli
        self.level_filter.blockSignals(True)
        current_selection = self.level_filter.currentText()
        self.level_filter.clear()
        self.level_filter.addItem("Tutti")
        for level in self.model.get_levels():
            self.level_filter.addItem(level)
        # Ripristina selezione se esiste ancora
        index = self.level_filter.findText(current_selection)
        if index >= 0:
            self.level_filter.setCurrentIndex(index)
        self.level_filter.blockSignals(False)

        self.update_table_info()

    def update_table_info(self):
        """Aggiorna info tabella"""
        count = len(self.model.nodes)
        self.table_info_label.setText(f"Nodi: {count}")

        # Aggiorna bounds
        if count > 0:
            bounds = self.model.get_bounds()
            bounds_text = (f"Range: X({bounds['x_min']:.2f}÷{bounds['x_max']:.2f}) "
                          f"Y({bounds['y_min']:.2f}÷{bounds['y_max']:.2f}) "
                          f"Z({bounds['z_min']:.2f}÷{bounds['z_max']:.2f})")
            self.bounds_label.setText(bounds_text)
        else:
            self.bounds_label.setText("")

    # === Reports & Export Methods ===

    def export_csv(self):
        """Export coordinate CSV"""
        if not self.model.nodes:
            QMessageBox.warning(self, "Export CSV", "Nessun nodo da esportare")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export CSV", "coordinate.csv",
            "CSV Files (*.csv);;All Files (*)"
        )

        if filename:
            try:
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['ID', 'X', 'Y', 'Z', 'Livello', 'Descrizione'])

                    for node in self.model.get_node_list():
                        writer.writerow([
                            node.id,
                            f"{node.x:.4f}",
                            f"{node.y:.4f}",
                            f"{node.z:.3f}",
                            node.level,
                            node.description
                        ])

                QMessageBox.information(self, "Export CSV",
                                      f"Esportati {len(self.model.nodes)} nodi in:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Export", f"Errore durante export:\n{e}")

    def export_excel(self):
        """Export coordinate Excel"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Export Excel",
                              "Openpyxl non disponibile.\nInstallare con: pip install openpyxl")
            return

        if not self.model.nodes:
            QMessageBox.warning(self, "Export Excel", "Nessun nodo da esportare")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", "coordinate.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if filename:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font as XLFont, PatternFill, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "Coordinate"

                # Header
                headers = ['ID', 'X (m)', 'Y (m)', 'Z (m)', 'Livello', 'Descrizione']
                ws.append(headers)

                # Stile header
                for cell in ws[1]:
                    cell.font = XLFont(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Dati
                for node in self.model.get_node_list():
                    ws.append([
                        node.id,
                        round(node.x, 4),
                        round(node.y, 4),
                        round(node.z, 3),
                        node.level,
                        node.description
                    ])

                # Formato numeri
                for row in ws.iter_rows(min_row=2, max_col=4):
                    for cell in row[1:4]:  # X, Y, Z
                        cell.number_format = '0.0000' if cell.column in [2, 3] else '0.000'
                        cell.alignment = Alignment(horizontal="right")

                # Auto-width colonne
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

                wb.save(filename)
                QMessageBox.information(self, "Export Excel",
                                      f"Esportati {len(self.model.nodes)} nodi in:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Export", f"Errore durante export:\n{e}")

    def export_pdf_report(self):
        """Export PDF report"""
        if not REPORTLAB_AVAILABLE:
            QMessageBox.warning(self, "Export PDF",
                              "Reportlab non disponibile.\nInstallare con: pip install reportlab")
            return

        if not self.model.nodes:
            QMessageBox.warning(self, "Export PDF", "Nessun nodo da esportare")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export PDF Report", "report.pdf",
            "PDF Files (*.pdf);;All Files (*)"
        )

        if filename:
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER

                # Crea documento
                doc = SimpleDocTemplate(filename, pagesize=A4)
                story = []
                styles = getSampleStyleSheet()

                # Titolo
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=18,
                    textColor=colors.HexColor('#366092'),
                    spaceAfter=30,
                    alignment=TA_CENTER
                )
                story.append(Paragraph("Report Coordinate Fili Fissi", title_style))
                story.append(Spacer(1, 0.5*cm))

                # Info progetto
                info_data = [
                    ['Progetto:', self.model.metadata.get('project_name', 'N/A')],
                    ['Località:', self.model.metadata.get('location', 'N/A')],
                    ['Autore:', self.model.metadata['author']],
                    ['Data:', datetime.now().strftime('%d/%m/%Y %H:%M')],
                    ['Totale Nodi:', str(len(self.model.nodes))],
                    ['Totale Edges:', str(len(self.model.edges))]
                ]

                info_table = Table(info_data, colWidths=[4*cm, 12*cm])
                info_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(info_table)
                story.append(Spacer(1, 1*cm))

                # Tabella coordinate
                coord_data = [['ID', 'X (m)', 'Y (m)', 'Z (m)', 'Livello', 'Descrizione']]
                for node in self.model.get_node_list()[:100]:  # Max 100 nodi
                    coord_data.append([
                        str(node.id),
                        f"{node.x:.4f}",
                        f"{node.y:.4f}",
                        f"{node.z:.3f}",
                        node.level,
                        node.description[:30] if node.description else ''
                    ])

                coord_table = Table(coord_data, colWidths=[1.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm, 5*cm])
                coord_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
                ]))
                story.append(coord_table)

                if len(self.model.nodes) > 100:
                    story.append(Spacer(1, 0.5*cm))
                    story.append(Paragraph(f"Nota: Mostrati solo i primi 100 nodi di {len(self.model.nodes)}", styles['Italic']))

                # Build PDF
                doc.build(story)

                QMessageBox.information(self, "Export PDF",
                                      f"Report PDF creato:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Errore Export", f"Errore durante export PDF:\n{e}")

    # === Import Methods ===

    def import_csv(self):
        """Import coordinate da file CSV"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "Import CSV", "",
            "CSV Files (*.csv);;All Files (*)"
        )

        if not filename:
            return

        try:
            imported = 0
            errors = []

            with open(filename, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)  # Skip header

                if not header:
                    QMessageBox.warning(self, "Import CSV", "File CSV vuoto")
                    return

                for row_num, row in enumerate(reader, start=2):
                    if len(row) < 3:  # Almeno X, Y, Z richiesti
                        errors.append(f"Riga {row_num}: dati insufficienti")
                        continue

                    try:
                        # ID opzionale (se non presente, auto-genera)
                        node_id = None
                        x_idx = 0
                        if len(row) >= 6:  # Formato completo: ID, X, Y, Z, Livello, Descrizione
                            try:
                                node_id = int(row[0])
                                x_idx = 1
                            except ValueError:
                                pass  # ID non valido, auto-genera

                        # Coordinate (obbligatorie)
                        x = float(row[x_idx])
                        y = float(row[x_idx + 1])
                        z = float(row[x_idx + 2])

                        # Dati opzionali
                        level = row[x_idx + 3] if len(row) > x_idx + 3 else ""
                        description = row[x_idx + 4] if len(row) > x_idx + 4 else ""

                        # Crea nodo con Command pattern
                        if node_id is not None and node_id in self.model.nodes:
                            errors.append(f"Riga {row_num}: ID {node_id} già esistente")
                            continue

                        node = Node(
                            id=node_id if node_id is not None else self.model.next_node_id,
                            x=x, y=y, z=z,
                            level=level,
                            description=description
                        )

                        cmd = CreateNodeCommand(self.model, node)
                        self.model.execute_command(cmd)
                        imported += 1

                    except (ValueError, IndexError) as e:
                        errors.append(f"Riga {row_num}: {str(e)}")
                        continue

            # Aggiorna tutte le viste
            self.refresh_all_views()
            self.update_status()
            self.update_undo_redo_actions()

            # Messaggio risultato
            msg = f"Importati {imported} nodi da CSV"
            if errors:
                msg += f"\n\nErrori ({len(errors)}):\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    msg += f"\n... e altri {len(errors) - 10} errori"
                QMessageBox.warning(self, "Import CSV", msg)
            else:
                QMessageBox.information(self, "Import CSV", msg)

        except Exception as e:
            QMessageBox.critical(self, "Errore Import", f"Errore durante import CSV:\n{e}")

    def import_excel(self):
        """Import coordinate da file Excel"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Import Excel",
                              "Openpyxl non disponibile.\nInstallare con: pip install openpyxl")
            return

        filename, _ = QFileDialog.getOpenFileName(
            self, "Import Excel", "",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not filename:
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename, data_only=True)

            # Cerca foglio "Coordinate" o usa il primo
            if "Coordinate" in wb.sheetnames:
                ws = wb["Coordinate"]
            else:
                ws = wb.active

            imported = 0
            errors = []

            # Salta header (riga 1)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue

                if len(row) < 3:
                    errors.append(f"Riga {row_num}: dati insufficienti")
                    continue

                try:
                    # ID opzionale
                    node_id = None
                    x_idx = 0
                    if len(row) >= 6:
                        try:
                            if row[0] is not None:
                                node_id = int(row[0])
                                x_idx = 1
                        except (ValueError, TypeError):
                            pass

                    # Coordinate
                    x = float(row[x_idx]) if row[x_idx] is not None else 0.0
                    y = float(row[x_idx + 1]) if row[x_idx + 1] is not None else 0.0
                    z = float(row[x_idx + 2]) if row[x_idx + 2] is not None else 0.0

                    # Dati opzionali
                    level = str(row[x_idx + 3]) if len(row) > x_idx + 3 and row[x_idx + 3] is not None else ""
                    description = str(row[x_idx + 4]) if len(row) > x_idx + 4 and row[x_idx + 4] is not None else ""

                    # Crea nodo
                    if node_id is not None and node_id in self.model.nodes:
                        errors.append(f"Riga {row_num}: ID {node_id} già esistente")
                        continue

                    node = Node(
                        id=node_id if node_id is not None else self.model.next_node_id,
                        x=x, y=y, z=z,
                        level=level,
                        description=description
                    )

                    cmd = CreateNodeCommand(self.model, node)
                    self.model.execute_command(cmd)
                    imported += 1

                except (ValueError, TypeError, IndexError) as e:
                    errors.append(f"Riga {row_num}: {str(e)}")
                    continue

            wb.close()

            # Aggiorna tutte le viste
            self.refresh_all_views()
            self.update_status()
            self.update_undo_redo_actions()

            # Messaggio risultato
            msg = f"Importati {imported} nodi da Excel"
            if errors:
                msg += f"\n\nErrori ({len(errors)}):\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    msg += f"\n... e altri {len(errors) - 10} errori"
                QMessageBox.warning(self, "Import Excel", msg)
            else:
                QMessageBox.information(self, "Import Excel", msg)

        except Exception as e:
            QMessageBox.critical(self, "Errore Import", f"Errore durante import Excel:\n{e}")

    def import_dxf(self):
        """Import coordinate da file DXF (AutoCAD)"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "Import DXF", "",
            "DXF Files (*.dxf);;All Files (*)"
        )

        if not filename:
            return

        try:
            imported = 0

            if EZDXF_AVAILABLE:
                # Import avanzato con ezdxf
                try:
                    doc = ezdxf.readfile(filename)
                    msp = doc.modelspace()

                    # Leggi POINT entities
                    for entity in msp.query('POINT'):
                        location = entity.dxf.location
                        x, y, z = location.x, location.y, location.z
                        level = entity.dxf.layer if hasattr(entity.dxf, 'layer') else ""
                        description = "DXF Point"

                        node = Node(
                            id=self.model.next_node_id,
                            x=x, y=y, z=z,
                            level=level,
                            description=description
                        )
                        cmd = CreateNodeCommand(self.model, node)
                        self.model.execute_command(cmd)
                        imported += 1

                    # Leggi TEXT entities come descrizioni di punti
                    text_entities = list(msp.query('TEXT'))
                    for text_entity in text_entities:
                        insert = text_entity.dxf.insert
                        text_content = text_entity.dxf.text
                        x, y, z = insert.x, insert.y, insert.z
                        level = text_entity.dxf.layer if hasattr(text_entity.dxf, 'layer') else ""

                        # Cerca se esiste già un punto molto vicino
                        point_exists = False
                        for node in self.model.nodes.values():
                            dist = math.sqrt((node.x - x)**2 + (node.y - y)**2 + (node.z - z)**2)
                            if dist < 0.01:  # Stesso punto
                                # Aggiorna descrizione se è ancora "DXF Point"
                                if not node.description or node.description == "DXF Point":
                                    node.description = text_content
                                point_exists = True
                                break

                        if not point_exists:
                            node = Node(
                                id=self.model.next_node_id,
                                x=x, y=y, z=z,
                                level=level,
                                description=text_content
                            )
                            cmd = CreateNodeCommand(self.model, node)
                            self.model.execute_command(cmd)
                            imported += 1

                    # Leggi INSERT entities (blocchi) come punti
                    for entity in msp.query('INSERT'):
                        insert = entity.dxf.insert
                        x, y, z = insert.x, insert.y, insert.z
                        block_name = entity.dxf.name
                        level = entity.dxf.layer if hasattr(entity.dxf, 'layer') else ""

                        node = Node(
                            id=self.model.next_node_id,
                            x=x, y=y, z=z,
                            level=level,
                            description=f"Block: {block_name}"
                        )
                        cmd = CreateNodeCommand(self.model, node)
                        self.model.execute_command(cmd)
                        imported += 1

                except Exception as e:
                    # Fallback a parser manuale
                    QMessageBox.warning(self, "Import DXF",
                                      f"Ezdxf fallito, uso parser manuale.\nErrore: {e}")
                    imported = self._import_dxf_manual(filename)

            else:
                # Parser DXF manuale (fallback)
                imported = self._import_dxf_manual(filename)

            # Aggiorna tutte le viste
            self.refresh_all_views()
            self.update_status()
            self.update_undo_redo_actions()

            msg = f"Importati {imported} nodi da DXF"
            if not EZDXF_AVAILABLE:
                msg += "\n\nNota: Ezdxf non disponibile, usato parser manuale.\nInstallare con: pip install ezdxf"

            QMessageBox.information(self, "Import DXF", msg)

        except Exception as e:
            QMessageBox.critical(self, "Errore Import", f"Errore durante import DXF:\n{e}")

    def _import_dxf_manual(self, filename: str) -> int:
        """Parser DXF manuale (fallback se ezdxf non disponibile)"""
        imported = 0

        with open(filename, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            # Cerca entità POINT
            if line == 'POINT':
                x, y, z = 0.0, 0.0, 0.0
                layer = ""
                # Leggi coordinate gruppo code 10, 20, 30
                j = i + 1
                while j < len(lines) and j < i + 50:
                    code = lines[j].strip()
                    if j + 1 < len(lines):
                        value = lines[j + 1].strip()
                        if code == '8':  # Layer
                            layer = value
                        elif code == '10':  # X
                            try:
                                x = float(value)
                            except:
                                pass
                        elif code == '20':  # Y
                            try:
                                y = float(value)
                            except:
                                pass
                        elif code == '30':  # Z
                            try:
                                z = float(value)
                            except:
                                pass
                        elif code == '0':  # Nuova entità
                            break
                    j += 2

                node = Node(
                    id=self.model.next_node_id,
                    x=x, y=y, z=z,
                    level=layer,
                    description="DXF Point"
                )
                cmd = CreateNodeCommand(self.model, node)
                self.model.execute_command(cmd)
                imported += 1

            # Cerca entità TEXT
            elif line == 'TEXT':
                x, y, z = 0.0, 0.0, 0.0
                text_content = ""
                layer = ""
                j = i + 1
                while j < len(lines) and j < i + 50:
                    code = lines[j].strip()
                    if j + 1 < len(lines):
                        value = lines[j + 1].strip()
                        if code == '8':  # Layer
                            layer = value
                        elif code == '10':  # X
                            try:
                                x = float(value)
                            except:
                                pass
                        elif code == '20':  # Y
                            try:
                                y = float(value)
                            except:
                                pass
                        elif code == '30':  # Z
                            try:
                                z = float(value)
                            except:
                                pass
                        elif code == '1':  # Testo
                            text_content = value
                        elif code == '0':  # Nuova entità
                            break
                    j += 2

                if text_content:
                    node = Node(
                        id=self.model.next_node_id,
                        x=x, y=y, z=z,
                        level=layer,
                        description=text_content
                    )
                    cmd = CreateNodeCommand(self.model, node)
                    self.model.execute_command(cmd)
                    imported += 1

            i += 1

        return imported

    def update_statistics(self):
        """Aggiorna statistiche progetto"""
        stats = []
        stats.append("=== STATISTICHE PROGETTO ===\n")

        # Conteggi base
        stats.append(f"Totale Nodi: {len(self.model.nodes)}")
        stats.append(f"Totale Edges: {len(self.model.edges)}")
        stats.append(f"Totale Annotazioni: {len(self.model.annotations)}\n")

        # Livelli
        levels = self.model.get_levels()
        if levels:
            stats.append(f"Livelli presenti ({len(levels)}): {', '.join(levels)}")
            for level in levels:
                level_nodes = self.model.get_nodes_by_level(level)
                stats.append(f"  - {level}: {len(level_nodes)} nodi")
        else:
            stats.append("Nessun livello definito")

        stats.append("")

        # Bounds
        if self.model.nodes:
            bounds = self.model.get_bounds()
            stats.append("=== COORDINATE ===")
            stats.append(f"X: {bounds['x_min']:.2f} ÷ {bounds['x_max']:.2f} (Δ={bounds['x_max']-bounds['x_min']:.2f}m)")
            stats.append(f"Y: {bounds['y_min']:.2f} ÷ {bounds['y_max']:.2f} (Δ={bounds['y_max']-bounds['y_min']:.2f}m)")
            stats.append(f"Z: {bounds['z_min']:.2f} ÷ {bounds['z_max']:.2f} (Δ={bounds['z_max']-bounds['z_min']:.2f}m)")
            stats.append("")

        # Edges stats
        if self.model.edges:
            stats.append("=== EDGES ===")
            lengths = []
            for edge in self.model.edges.values():
                n1 = self.model.nodes.get(edge.node1_id)
                n2 = self.model.nodes.get(edge.node2_id)
                if n1 and n2:
                    lengths.append(n1.distance_to(n2))

            if lengths:
                stats.append(f"Lunghezza minima: {min(lengths):.2f}m")
                stats.append(f"Lunghezza massima: {max(lengths):.2f}m")
                stats.append(f"Lunghezza media: {sum(lengths)/len(lengths):.2f}m")
                stats.append(f"Lunghezza totale: {sum(lengths):.2f}m")
            stats.append("")

        # Metadata
        stats.append("=== PROGETTO ===")
        stats.append(f"Nome: {self.model.metadata.get('project_name', 'N/A')}")
        stats.append(f"Località: {self.model.metadata.get('location', 'N/A')}")
        stats.append(f"Creato: {self.model.metadata.get('created', 'N/A')[:19]}")
        stats.append(f"Modificato: {self.model.metadata.get('modified', 'N/A')[:19]}")

        self.stats_text.setText('\n'.join(stats))

    def show_about(self):
        """Mostra informazioni"""
        about_text = f"""WIRE EDITOR UNIFIED v6.1

Applicazione integrata per gestione coordinate fili fissi
e design topologie strutturali.

Combina funzionalità di:
• wire_editor.py (input tabellare, import/export, reports)
• wire_editor1.py (canvas grafico, topologie)

Sviluppato per:
Arch. Michelangelo Bartolotta
Ordine Architetti Agrigento n. 1557

Librerie Opzionali:
Matplotlib: {"✓ Disponibile" if MATPLOTLIB_AVAILABLE else "✗ Non installato"}
Openpyxl: {"✓ Disponibile" if OPENPYXL_AVAILABLE else "✗ Non installato"}
Reportlab: {"✓ Disponibile" if REPORTLAB_AVAILABLE else "✗ Non installato"}
Ezdxf: {"✓ Disponibile" if EZDXF_AVAILABLE else "✗ Non installato"}

Versione: 6.1 (Import Features Complete - CSV, Excel, DXF)
"""
        QMessageBox.about(self, "About", about_text)


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName("Wire Editor Unified")
    app.setApplicationVersion("6.1")
    app.setOrganizationName("Arch. Michelangelo Bartolotta")

    # Stile applicazione
    app.setStyleSheet("""
        QTabWidget::pane {
            border: 1px solid #cccccc;
            border-radius: 3px;
        }
        QTabBar::tab {
            background: #e0e0e0;
            padding: 8px 16px;
            margin-right: 2px;
            border: 1px solid #cccccc;
            border-bottom: none;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
        }
        QTabBar::tab:selected {
            background: white;
            font-weight: bold;
        }
        QTabBar::tab:hover {
            background: #f0f0f0;
        }
    """)

    window = UnifiedMainWindow()
    window.show()

    sys.exit(app.exec_())
